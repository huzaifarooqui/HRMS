import os, math, secrets, json, re, calendar, subprocess, shutil
from io import BytesIO
from datetime import datetime,date,time,timedelta
from functools import wraps
from flask import Flask,render_template,request,redirect,url_for,session,flash,jsonify,send_from_directory,send_file,Response
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from config import Config
from database import init_pool,query,get_db
from schema import setup_database
import digilocker as dl
from itsdangerous import URLSafeSerializer, BadSignature
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.graphics.shapes import Drawing
from reportlab.graphics import renderSVG, renderPDF
from offer_letter import build_offer_letter_pdf
from appointment_letter import build_appointment_letter_pdf
from termination_letter import build_termination_letter_pdf
from promotion_letter import build_promotion_letter_pdf
from pay_slip import build_pay_slip_pdf
from experience_letter import build_experience_letter_pdf
from warning_letter import build_warning_letter_pdf
from relieving_letter import build_relieving_letter_pdf
from pip_letter import build_pip_letter_pdf
from digital_id_pdf_renderer import build_digital_id_pdf, file_to_data_uri, svg_to_data_uri

app=Flask(__name__); app.config.from_object(Config)
# Role-aware inactivity protection; browser keepalive prevents logout while the user is actively working.
# Cookie lifetime is slightly above the admin inactivity limit; role-specific enforcement is below.
app.config['PERMANENT_SESSION_LIFETIME']=timedelta(minutes=12)
app.config['SESSION_REFRESH_EACH_REQUEST']=True
# Persistent uploads: survive future code/build replacements.
PERSISTENT_UPLOAD_ROOT=app.config.get('UPLOAD_ROOT') or os.path.dirname(app.config['UPLOAD_FOLDER'])
COMPANY_UPLOAD=os.path.join(PERSISTENT_UPLOAD_ROOT,'company')
DOCUMENT_UPLOAD=os.path.join(PERSISTENT_UPLOAD_ROOT,'documents')
os.makedirs(app.config['UPLOAD_FOLDER'],exist_ok=True)
os.makedirs(COMPANY_UPLOAD,exist_ok=True)
os.makedirs(DOCUMENT_UPLOAD,exist_ok=True)

def migrate_legacy_uploads_once():
    """Copy any legacy Server/uploads files into persistent storage, without overwriting."""
    import shutil
    legacy_root=os.path.join(os.path.dirname(__file__),'uploads')
    targets={
        'employees':app.config['UPLOAD_FOLDER'],
        'company':COMPANY_UPLOAD,
        'documents':DOCUMENT_UPLOAD,
    }
    if not os.path.isdir(legacy_root):
        return
    for bucket,target in targets.items():
        source=os.path.join(legacy_root,bucket)
        if not os.path.isdir(source):
            continue
        os.makedirs(target,exist_ok=True)
        for name in os.listdir(source):
            src=os.path.join(source,name)
            dst=os.path.join(target,name)
            if os.path.isfile(src) and not os.path.exists(dst):
                try: shutil.copy2(src,dst)
                except OSError: pass

migrate_legacy_uploads_once()
init_pool(app)
with app.app_context():
    setup_database()
    # FINAL attendance rule: Leave and Absent are one attendance status.
    # Preserve leave-request workflow, but attendance register/payroll uses Absent only.
    try:
        query("UPDATE attendance SET status='Absent' WHERE status='Leave'",commit=True)
    except Exception:
        pass

ALLOWED={'png','jpg','jpeg','webp'}
DOC_ALLOWED={'pdf','png','jpg','jpeg','doc','docx'}
BLOOD_GROUPS=['A+','A-','B+','B-','AB+','AB-','O+','O-']
EMPLOYEE_DOCUMENT_TYPES=[
'Aadhaar Card','PAN Card','Driving Licence','Voter ID','Passport',
'10th Marksheet / Certificate','12th Marksheet / Certificate','Graduation / Qualification',
'Resume / CV','Bank Document / Cancelled Cheque','Address Proof','Police Verification',
'Experience Letter','Relieving Letter','Previous Employment Document','Offer Letter',
'Appointment Letter','Warning Letter','Performance Improvement Plan (PIP)','Promotion Letter','Termination Letter','Pay Slip','Other'
]
ATTENDANCE_NO_CHECKIN_CUTOFF=time(14,0)
ATTENDANCE_EARLY_CHECKOUT_CUTOFF=time(16,0)
INDIA_STATES=['Andaman and Nicobar Islands','Andhra Pradesh','Arunachal Pradesh','Assam','Bihar','Chandigarh','Chhattisgarh','Dadra and Nagar Haveli and Daman and Diu','Delhi','Goa','Gujarat','Haryana','Himachal Pradesh','Jammu and Kashmir','Jharkhand','Karnataka','Kerala','Ladakh','Lakshadweep','Madhya Pradesh','Maharashtra','Manipur','Meghalaya','Mizoram','Nagaland','Odisha','Puducherry','Punjab','Rajasthan','Sikkim','Tamil Nadu','Telangana','Tripura','Uttar Pradesh','Uttarakhand','West Bengal']

def admin_required(f):
    @wraps(f)
    def w(*a,**k):
        aid=session.get('admin_id')
        if not aid: return redirect(url_for('admin_login'))
        adm=query("SELECT id,is_active FROM admins WHERE id=%s",(aid,),one=True)
        if not adm or int(adm.get('is_active') if adm.get('is_active') is not None else 1)!=1:
            session.clear(); flash('Administrator access has been disabled.','warning'); return redirect(url_for('admin_login'))
        return f(*a,**k)
    return w

ADMIN_ROLES=('Super Admin','Owner','Manager','HR')

def current_admin():
    if not session.get('admin_id'): return None
    return query("SELECT * FROM admins WHERE id=%s AND COALESCE(is_active,1)=1",(session['admin_id'],),one=True)

def current_admin_role():
    adm=current_admin()
    return (adm.get('role') if adm else None) or 'Super Admin'

def role_required(*roles):
    def deco(f):
        @wraps(f)
        def w(*a,**k):
            if not session.get('admin_id'): return redirect(url_for('admin_login'))
            if current_admin_role() not in roles:
                flash('You do not have permission for this action.','danger')
                return redirect(request.referrer or url_for('admin_dashboard'))
            return f(*a,**k)
        return w
    return deco

def can_manage_everything(): return current_admin_role()=='Super Admin'

def pending_approval_count():
    role=current_admin_role()
    if role=='Super Admin':
        row=query("SELECT COUNT(*) c FROM approval_requests WHERE status='Pending'",one=True)
    elif role in ('HR','Manager','Owner'):
        row=query("SELECT COUNT(*) c FROM approval_requests WHERE status='Pending' AND current_stage=%s",(role,),one=True)
    else: row={'c':0}
    return int((row or {}).get('c') or 0)

def employee_required(f):
    @wraps(f)
    def w(*a,**k):
        if not session.get('employee_id'): return redirect(url_for('employee_login'))
        return f(*a,**k)
    return w

def settings(): return query("SELECT * FROM company_settings WHERE id=1",one=True) or {}

def as_time(value, default):
    if value is None: return default
    if isinstance(value,time): return value
    if isinstance(value,timedelta):
        total_seconds=int(value.total_seconds())%(24*60*60); hours,remainder=divmod(total_seconds,3600); minutes,seconds=divmod(remainder,60); return time(hours,minutes,seconds)
    if isinstance(value,str):
        try:
            parts=value.strip().split(':'); return time(int(parts[0]),int(parts[1]),int(float(parts[2])) if len(parts)>2 else 0)
        except (ValueError,TypeError,IndexError): return default
    return default

def time_value(v):
    if v is None:return ''
    if isinstance(v,timedelta):return as_time(v,time(0,0)).strftime('%H:%M')
    if hasattr(v,'strftime'):return v.strftime('%H:%M')
    text=str(v);return text[:5] if len(text)>=5 else text
app.jinja_env.filters['time_value']=time_value

def minutes_to_hm(n):
    n=int(n or 0);return f"{n//60}h {n%60:02d}m"
app.jinja_env.filters['hm']=minutes_to_hm

def distance_m(lat1,lng1,lat2,lng2):
    r=6371000;p1=math.radians(float(lat1));p2=math.radians(float(lat2));dp=math.radians(float(lat2)-float(lat1));dl=math.radians(float(lng2)-float(lng1));a=math.sin(dp/2)**2+math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2;return 2*r*math.asin(math.sqrt(a))

def clean_location(form,name):
    value=(form.get(name) or '').strip()
    custom=(form.get(name+'_custom') or '').strip()
    return custom if value=='__other__' and custom else value

def unique_login_id(first_name,mother_name,joining_date):
    """First initial + mother's first-name initial + joining YYMMDD + random 2 digits."""
    first=(first_name or '').strip(); mother=(mother_name or '').strip()
    if not first or not mother or not joining_date: raise ValueError("First name, mother's name and joining date are required to generate Login ID.")
    try: join=datetime.strptime(str(joining_date),'%Y-%m-%d').date()
    except ValueError: raise ValueError('Joining date is invalid.')
    prefix=(first[0]+mother.split()[0][0]+join.strftime('%y%m%d')).upper()
    for _ in range(150):
        code=prefix+f"{secrets.randbelow(100):02d}"
        if not query("SELECT id FROM employees WHERE login_id=%s",(code,),one=True): return code
    raise ValueError('Could not generate a unique Login ID. Please try again.')

def parse_dt(att_date,clock_value):
    if not clock_value:return None
    return datetime.combine(att_date,datetime.strptime(clock_value,'%H:%M').time())

def attendance_metrics(att_date,check_in,check_out):
    s=settings(); report=datetime.combine(att_date,as_time(s.get('last_reporting'),time(9,45))); end=datetime.combine(att_date,as_time(s.get('office_end'),time(18,30)))
    late=max(0,int((check_in-report).total_seconds()//60)) if check_in else 0
    working=max(0,int((check_out-check_in).total_seconds()//60)) if check_in and check_out else 0
    overtime=max(0,int((check_out-end).total_seconds()//60)) if check_out else 0
    early=max(0,int((end-check_out).total_seconds()//60)) if check_out else 0
    return working,late,overtime,early



def _session_timeout_seconds():
    if session.get('employee_id'):
        return 5 * 60
    if session.get('admin_id'):
        return 10 * 60
    return None

@app.before_request
def enforce_role_inactivity_timeout():
    """True inactivity timeout: employee 5 min, admin 10 min.

    Browser activity pings refresh _last_activity, so users actively typing in long
    forms are not logged out mid-save.
    """
    limit=_session_timeout_seconds()
    if not limit:
        return None
    now=int(datetime.now().timestamp())
    last=session.get('_last_activity')
    if last is not None and now-int(last)>limit:
        was_employee=bool(session.get('employee_id'))
        session.clear()
        if request.endpoint=='session_keepalive':
            return jsonify({'success':False,'expired':True}),401
        return redirect(url_for('employee_login' if was_employee else 'admin_login',reason='timeout'))
    session['_last_activity']=now
    return None

@app.route('/session/keepalive',methods=['POST'])
def session_keepalive():
    if not session.get('admin_id') and not session.get('employee_id'):
        return jsonify({'success':False,'expired':True}),401
    session['_last_activity']=int(datetime.now().timestamp())
    return jsonify({'success':True})


def audit_log(action,entity_type,entity_id=None,details=None,employee_id=None):
    try:
        query("""INSERT INTO system_audit_log(admin_id,employee_id,action,entity_type,entity_id,details)
                 VALUES(%s,%s,%s,%s,%s,%s)""",
              (session.get('admin_id'),employee_id,action,entity_type,
               str(entity_id) if entity_id is not None else None,
               json.dumps(details,default=str,ensure_ascii=False) if isinstance(details,(dict,list)) else (str(details) if details else None)),
              commit=True)
    except Exception:
        pass

def notify_employee(employee_id,title,message,notification_type='General'):
    try:
        query("""INSERT INTO employee_notifications(employee_id,title,message,notification_type)
                 VALUES(%s,%s,%s,%s)""",
              (employee_id,title[:160],message[:500],notification_type[:50]),commit=True)
    except Exception:
        pass

def month_is_locked(month_key):
    row=query("SELECT is_locked FROM attendance_month_locks WHERE month_key=%s",(month_key,),one=True)
    return bool(row and row.get('is_locked'))

def payroll_is_finalized(month_key):
    row=query("SELECT is_finalized FROM payroll_locks WHERE month_key=%s",(month_key,),one=True)
    return bool(row and row.get('is_finalized'))

def profile_completion(emp):
    fields=['photo','first_name','mother_name','dob','gender','email','phone','address','city','district','state','pincode',
            'aadhar','pan','department_id','designation','joining_date','salary','blood_group']
    done=sum(1 for f in fields if emp.get(f) not in (None,''))
    return round((done/len(fields))*100) if fields else 100

def attendance_anomaly_count(month_key=None):
    month_key=month_key or date.today().strftime('%Y-%m')
    row=query("""SELECT COUNT(*) c FROM attendance
                 WHERE DATE_FORMAT(attendance_date,'%Y-%m')=%s
                 AND ((check_in IS NOT NULL AND check_out IS NULL AND attendance_date<CURDATE())
                      OR (check_out IS NOT NULL AND check_in IS NULL))""",(month_key,),one=True)
    return int((row or {}).get('c') or 0)

def sync_company_attendance_rules(day=None):
    now=datetime.now()
    day=day or now.date()
    employees = query("""
        SELECT id, joining_date, status, inactive_from
        FROM employees
        WHERE status='Active'
        AND (
                inactive_from IS NULL
                OR inactive_from > %s
            )
    """, (day,))

    # Highest-priority company exception: first date of every month is auto Present.
    if day.day==1:
        for e in employees:
            if e.get('joining_date') and e['joining_date']>day: continue
            row=query("SELECT id FROM attendance WHERE employee_id=%s AND attendance_date=%s",(e['id'],day),one=True)
            if not row:
                query("""INSERT INTO attendance(employee_id,attendance_date,status,working_minutes,late_minutes,
                         overtime_minutes,early_exit_minutes,remarks)
                         VALUES(%s,%s,'Present',0,0,0,0,'Auto Present - First day of month')""",
                      (e['id'],day),commit=True)
        return

    # Sunday paid day / sandwich rule already finalized earlier.
    if day.weekday()==6:
        try: sync_sunday_salary_days(day,day)
        except Exception: pass
        return

    # At/after 2 PM, no check-in means company leave. Attendance/reporting keeps
    # the previously approved merged Leave/Absent behavior by storing Absent.
    if day!=now.date() or now.time()<ATTENDANCE_NO_CHECKIN_CUTOFF:
        return
    for e in employees:
        if e.get('joining_date') and e['joining_date']>day: continue
        existing=query("SELECT id FROM attendance WHERE employee_id=%s AND attendance_date=%s",(e['id'],day),one=True)
        if existing: continue
        query("""INSERT INTO attendance(employee_id,attendance_date,status,working_minutes,late_minutes,
                 overtime_minutes,early_exit_minutes,remarks)
                 VALUES(%s,%s,'Absent',0,0,0,0,'Auto Leave: no check-in by 2:00 PM')""",
              (e['id'],day),commit=True)

def dashboard_birthdays(include_admins=True):
    """Return birthdays in the next 30 days, with today's birthdays first."""
    today=date.today()
    people=[]
    emps=query("SELECT id,first_name,last_name,dob FROM employees WHERE status='Active' AND dob IS NOT NULL")
    for e in emps:
        people.append({'kind':'Employee','id':e['id'],'name':((e.get('first_name') or '')+' '+(e.get('last_name') or '')).strip(),'dob':e['dob']})
    if include_admins:
        try:
            admins=query("SELECT id,full_name,username,dob,role FROM admins WHERE COALESCE(is_active,1)=1 AND dob IS NOT NULL")
            for a in admins:
                people.append({'kind':'Admin','id':a['id'],'name':(a.get('full_name') or a.get('username') or 'Administrator').strip(),'dob':a['dob'],'role':a.get('role') or 'Administrator'})
        except Exception:
            pass
    out=[]
    for p in people:
        try:
            b=date(today.year,p['dob'].month,p['dob'].day)
        except ValueError:
            b=date(today.year,2,28)
        if b < today:
            try: b=date(today.year+1,p['dob'].month,p['dob'].day)
            except ValueError: b=date(today.year+1,2,28)
        delta=(b-today).days
        if 0 <= delta <= 30:
            p['birthday_date']=b; p['days_until']=delta; out.append(p)
    out.sort(key=lambda p:(p['days_until'], 0 if p['kind']=='Employee' else 1, p['name'].lower()))
    return out

@app.context_processor
def ctx():
    unread=0
    if session.get('employee_id'):
        try:
            r=query("SELECT COUNT(*) c FROM employee_notifications WHERE employee_id=%s AND is_read=0",(session['employee_id'],),one=True)
            unread=int((r or {}).get('c') or 0)
        except Exception: unread=0
    role=current_admin_role() if session.get('admin_id') else None
    approvals=0
    if session.get('admin_id'):
        try: approvals=pending_approval_count()
        except Exception: approvals=0
    return {'company':settings(),'today':date.today(),'blood_groups':BLOOD_GROUPS,'india_states':INDIA_STATES,
            'employee_document_types':EMPLOYEE_DOCUMENT_TYPES,'unread_notifications':unread,
            'admin_role':role,'admin_roles':ADMIN_ROLES,'pending_approvals':approvals,'expiring_documents':int((query("SELECT COUNT(*) c FROM employee_documents WHERE COALESCE(document_status,'Active')<>'Deleted' AND expiry_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(),INTERVAL 30 DAY)",one=True) or {}).get('c') or 0) if session.get('admin_id') else 0,'open_hr_cases':int((query("SELECT COUNT(*) c FROM hr_cases WHERE status='Open'",one=True) or {}).get('c') or 0) if session.get('admin_id') else 0}

@app.route('/')
def home(): return render_template('home.html')
@app.route('/uploads/employees/<path:name>')
def uploaded(name): return send_from_directory(app.config['UPLOAD_FOLDER'],name)
@app.route('/uploads/company/<path:name>')
def company_uploaded(name): return send_from_directory(COMPANY_UPLOAD,name)
@app.route('/uploads/documents/<path:name>')
def document_uploaded(name):
    if not session.get('admin_id') and not session.get('employee_id'): return redirect(url_for('home'))
    return send_from_directory(DOCUMENT_UPLOAD,name,as_attachment=False)

@app.route('/admin/login',methods=['GET','POST'])
def admin_login():
    if request.method=='GET' and request.args.get('reason')=='timeout': flash('Your session expired after extended inactivity. Please sign in again.','warning')
    if request.method=='POST':
        u=request.form.get('username','').strip();p=request.form.get('password','');adm=query("SELECT * FROM admins WHERE username=%s AND COALESCE(is_active,1)=1",(u,),one=True)
        valid=False
        if adm:
            stored=adm.get('password') or ''
            try: valid=check_password_hash(stored,p) if stored.startswith(('pbkdf2:','scrypt:')) else stored==p
            except Exception: valid=stored==p
        if adm and valid:
            if not stored.startswith(('pbkdf2:','scrypt:')):
                try: query("UPDATE admins SET password=%s WHERE id=%s",(generate_password_hash(p),adm['id']),commit=True)
                except Exception: pass
            session.clear();session.permanent=True;session['admin_id']=adm['id'];session['admin_name']=adm['full_name'] or adm['username'];session['admin_role']=adm.get('role') or 'Super Admin';session['_last_activity']=int(datetime.now().timestamp());query("UPDATE admins SET last_login=NOW() WHERE id=%s",(adm['id'],),commit=True); query("INSERT INTO admin_login_history(admin_id,ip_address,user_agent) VALUES(%s,%s,%s)",(adm['id'],request.remote_addr,(request.headers.get('User-Agent') or '')[:500]),commit=True); return redirect(url_for('admin_dashboard'))
        flash('Invalid username or password.','danger')
    return render_template('admin_login.html')
@app.route('/admin/logout')
def admin_logout():session.clear();return redirect(url_for('admin_login'))

@app.route('/admin/profile')
@admin_required
def admin_profile():
    admin=query("SELECT id,username,full_name,dob,created_at,last_login FROM admins WHERE id=%s",(session['admin_id'],),one=True)
    if not admin:session.clear();flash('Administrator account could not be found.','danger');return redirect(url_for('admin_login'))
    return render_template('admin_profile.html',admin=admin)


@app.route('/admin/access',methods=['GET','POST'])
@admin_required
@role_required('Super Admin')
def admin_access_control():
    if request.method=='POST':
        action=request.form.get('action') or 'create'
        try:
            if action=='create':
                username=(request.form.get('username') or '').strip()
                full_name=(request.form.get('full_name') or '').strip()
                role=(request.form.get('role') or '').strip()
                password=request.form.get('password') or ''
                if not username or not full_name or role not in ADMIN_ROLES or len(password)<8:
                    raise ValueError('Full name, unique username, valid role and password of at least 8 characters are required.')
                aid=query("INSERT INTO admins(username,password,full_name,dob,must_change_password,role,is_active) VALUES(%s,%s,%s,NULLIF(%s,''),1,%s,1)",(username,generate_password_hash(password),full_name,request.form.get('dob',''),role),commit=True)
                audit_log('CREATE_ADMIN','Admin',aid,{'username':username,'role':role})
                flash(f'{role} account created successfully.','success')
            elif action=='update':
                aid=int(request.form.get('admin_id') or 0)
                target=query("SELECT * FROM admins WHERE id=%s",(aid,),one=True)
                if not target: raise ValueError('Administrator not found.')
                role=(request.form.get('role') or target.get('role') or 'HR').strip()
                active=1 if request.form.get('is_active')=='1' else 0
                if role not in ADMIN_ROLES: raise ValueError('Invalid role.')
                if aid==session['admin_id'] and not active: raise ValueError('You cannot deactivate your own account.')
                if target.get('role')=='Super Admin' and aid!=session['admin_id'] and role!='Super Admin':
                    pass
                query("UPDATE admins SET full_name=%s,dob=NULLIF(%s,''),role=%s,is_active=%s WHERE id=%s",((request.form.get('full_name') or target.get('full_name') or '').strip(),request.form.get('dob',''),role,active,aid),commit=True)
                new_password=request.form.get('new_password') or ''
                if new_password:
                    if len(new_password)<8: raise ValueError('New password must be at least 8 characters.')
                    query("UPDATE admins SET password=%s,must_change_password=1 WHERE id=%s",(generate_password_hash(new_password),aid),commit=True)
                audit_log('UPDATE_ADMIN','Admin',aid,{'role':role,'active':active})
                flash('Administrator access updated.','success')
        except Exception as e: flash(f'Unable to update administrator access: {e}','danger')
        return redirect(url_for('admin_access_control'))
    admins=query("SELECT id,username,full_name,dob,role,is_active,created_at,last_login FROM admins ORDER BY FIELD(role,'Super Admin','Owner','Manager','HR'),full_name")
    return render_template('admin_access_control.html',admins=admins)

@app.route('/admin/approvals')
@admin_required
def admin_approvals():
    role=current_admin_role(); params=[]; where="ar.status='Pending'"
    if role!='Super Admin':
        where += " AND ar.current_stage=%s"; params.append(role)
    rows=query(f"""SELECT ar.*,a.full_name submitted_name,a.role submitted_role
                    FROM approval_requests ar JOIN admins a ON a.id=ar.submitted_by
                    WHERE {where} ORDER BY ar.created_at ASC""",tuple(params))
    history=query("""SELECT ar.*,a.full_name submitted_name FROM approval_requests ar JOIN admins a ON a.id=ar.submitted_by
                     WHERE ar.status<>'Pending' ORDER BY ar.updated_at DESC LIMIT 40""")
    return render_template('admin_approvals.html',rows=rows,history=history)

def finalize_leave_approval(lid,remark=''):
    query("UPDATE leave_requests SET status='Approved',admin_remark=%s,reviewed_by=%s,reviewed_at=NOW() WHERE id=%s",(remark,session['admin_id'],lid),commit=True)
    lr=query("SELECT employee_id,start_date,end_date FROM leave_requests WHERE id=%s",(lid,),one=True)
    if lr:
        apply_approved_leave_as_absent(lid)
        sync_sunday_salary_days(lr['start_date']-timedelta(days=1),lr['end_date']+timedelta(days=1))
        notify_employee(lr['employee_id'],'Leave Approved','Your leave request has received final approval.','Leave')
        audit_log('LEAVE_APPROVED_FINAL','LeaveRequest',lid,{'remark':remark},lr['employee_id'])

def _coerce_db_time(value):
    """Normalize MySQL TIME values returned as time, timedelta, or text."""
    if value is None:
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, timedelta):
        total_seconds = int(value.total_seconds()) % 86400
        return time(total_seconds // 3600, (total_seconds % 3600) // 60, total_seconds % 60)
    if isinstance(value, str):
        text = value.strip()
        for fmt in ('%H:%M:%S', '%H:%M'):
            try:
                return datetime.strptime(text, fmt).time()
            except ValueError:
                pass
    raise TypeError(f"Unsupported database TIME value: {type(value).__name__}")


def apply_approval_record(row):
    payload=json.loads(row.get('payload_json') or '{}')
    if row.get('request_type')=='Employee Profile Update':
        eid=int(row['entity_id']); apply_employee_profile_payload(eid,payload)
        audit_log('APPROVE_EMPLOYEE_UPDATE','Employee',eid,{'approval_id':row['id']},eid)
    elif row.get('request_type')=='Leave Approval':
        finalize_leave_approval(int(row['entity_id']),payload.get('remark') or '')
    elif row.get('request_type')=='Attendance Regularization':
        rid=int(row['entity_id']); rr=query("SELECT * FROM attendance_regularization WHERE id=%s",(rid,),one=True)
        if rr:
            existing=query("SELECT id FROM attendance WHERE employee_id=%s AND attendance_date=%s",(rr['employee_id'],rr['attendance_date']),one=True)
            ci=(datetime.combine(rr['attendance_date'],_coerce_db_time(rr['requested_check_in'])) if rr.get('requested_check_in') else None); co=(datetime.combine(rr['attendance_date'],_coerce_db_time(rr['requested_check_out'])) if rr.get('requested_check_out') else None)
        if existing:
            query("""
                UPDATE attendance
                SET
                    check_in=COALESCE(%s,check_in),
                    check_out=COALESCE(%s,check_out),
                    status=COALESCE(NULLIF(%s,''),status),
                    remarks=%s
                WHERE id=%s
            """,(
                ci,
                co,
                rr.get('requested_status') or '',
                'Regularized: ' + rr['reason'],
                existing['id']
            ),commit=True)
        else:
            requested_status = rr.get('requested_status') or 'Present'

            query("""
                INSERT INTO attendance(
                    employee_id,
                    attendance_date,
                    check_in,
                    check_out,
                    status,
                    remarks
                )
                VALUES(%s,%s,%s,%s,%s,%s)
            """,(
                rr['employee_id'],
                rr['attendance_date'],
                ci,
                co,
                requested_status,
                'Regularized: ' + rr['reason']
            ),commit=True)
            query("UPDATE attendance_regularization SET status='Approved' WHERE id=%s",(rid,),commit=True); notify_employee(rr['employee_id'],'Attendance regularized',f"Your attendance correction for {rr['attendance_date']} was approved.",'Attendance')

@app.post('/admin/approvals/<int:approval_id>/decision')
@admin_required
def admin_approval_decision(approval_id):
    row=query("SELECT * FROM approval_requests WHERE id=%s",(approval_id,),one=True)
    if not row or row.get('status')!='Pending':
        flash('Approval request is no longer pending.','warning'); return redirect(url_for('admin_approvals'))
    role=current_admin_role(); decision=request.form.get('decision'); note=(request.form.get('note') or '').strip()[:500]
    if role!='Super Admin' and role!=row.get('current_stage'):
        flash('This approval is assigned to another approval level.','danger'); return redirect(url_for('admin_approvals'))
    if decision=='reject':
        query("UPDATE approval_requests SET status='Rejected',current_stage='Completed',rejected_by=%s,rejected_at=NOW(),rejection_note=%s WHERE id=%s",(session['admin_id'],note or None,approval_id),commit=True)
        if row.get('request_type')=='Leave Approval':
            query("UPDATE leave_requests SET status='Rejected',admin_remark=%s,reviewed_by=%s,reviewed_at=NOW() WHERE id=%s",(note or 'Rejected during approval workflow',session['admin_id'],int(row['entity_id'])),commit=True)
        audit_log('REJECT_APPROVAL','ApprovalRequest',approval_id,{'note':note})
        flash('Request rejected and workflow closed.','warning'); return redirect(url_for('admin_approvals'))
    if decision!='approve':
        flash('Choose Approve or Reject.','warning'); return redirect(url_for('admin_approvals'))
    if role=='Super Admin':
        apply_approval_record(row)
        query("UPDATE approval_requests SET status='Approved',current_stage='Completed',owner_by=%s,owner_at=NOW(),owner_note=%s WHERE id=%s",(session['admin_id'],note or 'Super Admin final approval',approval_id),commit=True)
        audit_log('SUPERADMIN_APPROVE','ApprovalRequest',approval_id,{'note':note})
        flash('Request approved and applied by Super Admin.','success'); return redirect(url_for('admin_approvals'))
    if role=='Manager':
        query("UPDATE approval_requests SET manager_by=%s,manager_at=NOW(),manager_note=%s,current_stage='Owner',stage_entered_at=NOW() WHERE id=%s",(session['admin_id'],note or None,approval_id),commit=True)
        flash('Manager approval recorded. Owner final approval is now required.','success')
    elif role=='Owner':
        apply_approval_record(row)
        query("UPDATE approval_requests SET owner_by=%s,owner_at=NOW(),owner_note=%s,status='Approved',current_stage='Completed' WHERE id=%s",(session['admin_id'],note or None,approval_id),commit=True)
        flash('Owner final approval completed and changes applied.','success')
    elif role=='HR':
        query("UPDATE approval_requests SET hr_by=%s,hr_at=NOW(),hr_note=%s,current_stage='Manager',stage_entered_at=NOW() WHERE id=%s",(session['admin_id'],note or None,approval_id),commit=True)
        flash('HR approval recorded. Manager approval is now required.','success')
    audit_log('APPROVAL_DECISION','ApprovalRequest',approval_id,{'role':role,'decision':'approve','note':note})
    return redirect(url_for('admin_approvals'))

@app.route('/api/login',methods=['GET','POST'])
def employee_login():
    if request.method=='GET':
        if request.args.get('reason')=='timeout':flash('Your session expired after extended inactivity. Please sign in again.','warning')
        return render_template('employee_login.html')
    data=request.form if request.form else request.get_json(silent=True) or {};code=data.get('login_id','').strip();pwd=data.get('password','').strip();emp=query("SELECT * FROM employees WHERE login_id=%s AND status='Active'",(code,),one=True);valid=emp and ((not emp.get('password')) or emp.get('password')==pwd)
    if valid:
        session.clear();session.permanent=True;session['employee_id']=emp['id'];session['employee_name']=(emp['first_name']+' '+(emp['last_name'] or '')).strip();session['_last_activity']=int(datetime.now().timestamp());return redirect(url_for('employee_dashboard'))
    flash('Invalid employee login ID or password.','danger');return render_template('employee_login.html'),401
@app.route('/employee/logout')
def employee_logout():session.clear();return redirect(url_for('employee_login'))

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    try: sync_company_attendance_rules()
    except Exception: pass
    counts=query("""SELECT (SELECT COUNT(*) FROM employees WHERE status='Active') employees,(SELECT COUNT(*) FROM attendance WHERE attendance_date=CURDATE() AND status IN('Present','Late','Half Day')) present,(SELECT COUNT(*) FROM attendance WHERE attendance_date=CURDATE() AND status='Late') late,(SELECT COUNT(*) FROM leave_requests WHERE status='Pending') pending""",one=True)
    recent=query("""SELECT a.*,CONCAT(e.first_name,' ',COALESCE(e.last_name,'')) employee_name,e.login_id,e.designation FROM attendance a JOIN employees e ON e.id=a.employee_id WHERE a.attendance_date=CURDATE() ORDER BY a.check_in DESC LIMIT 12""")
    upcoming=query("SELECT * FROM holidays WHERE status='Active' AND holiday_date>=CURDATE() ORDER BY holiday_date LIMIT 5")
    expiring_docs=query("""SELECT COUNT(*) c FROM employee_documents
                              WHERE COALESCE(document_status,'Active')<>'Deleted'
                              AND expiry_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(),INTERVAL 30 DAY)""",one=True)
    feedback_new=query("SELECT COUNT(*) c FROM employee_feedback WHERE status='New'",one=True)
    alerts={'expiring_docs':int((expiring_docs or {}).get('c') or 0),
            'attendance_anomalies':attendance_anomaly_count(date.today().strftime('%Y-%m')),
            'feedback_new':int((feedback_new or {}).get('c') or 0),
            'pending_leaves':int((counts or {}).get('pending') or 0),
            'pending_approvals':pending_approval_count()}
    allocation_log=os.path.join(os.path.dirname(__file__),'allocation_sync.log')
    health={'database':'Connected','allocation_log':os.path.exists(allocation_log),
            'backup_dir':os.path.isdir(os.path.join(app.config.get('DATA_ROOT',os.path.dirname(__file__)),'backups'))}
    return render_template('admin_dashboard.html',counts=counts,recent=recent,upcoming=upcoming,
                           now_hour=datetime.now().hour,alerts=alerts,health=health,birthdays=dashboard_birthdays(True))


@app.route('/admin/attendance-calendar')
@admin_required
def admin_attendance_calendar():
    month=(request.args.get('month') or date.today().strftime('%Y-%m'))[:7]
    employee_id=(request.args.get('employee_id') or '').strip()
    try:
        year,mon=map(int,month.split('-'))
        first_day=date(year,mon,1)
        last_day=date(year,mon,calendar.monthrange(year,mon)[1])
    except Exception:
        today=date.today()
        year,mon=today.year,today.month
        month=today.strftime('%Y-%m')
        first_day=date(year,mon,1)
        last_day=date(year,mon,calendar.monthrange(year,mon)[1])

    employees=query("""SELECT id,login_id,first_name,last_name
                       FROM employees WHERE status='Active'
                       ORDER BY first_name,last_name""")

    params=[first_day,last_day]
    where="a.attendance_date BETWEEN %s AND %s"
    if employee_id.isdigit():
        where += " AND a.employee_id=%s"
        params.append(int(employee_id))

    rows=query(f"""SELECT a.employee_id,a.attendance_date,a.check_in,a.check_out,
                          a.status,a.working_minutes,a.late_minutes,a.overtime_minutes,
                          CONCAT(e.first_name,' ',COALESCE(e.last_name,'')) employee_name,
                          e.login_id
                   FROM attendance a
                   JOIN employees e ON e.id=a.employee_id
                   WHERE {where}
                   ORDER BY a.attendance_date""",tuple(params))

    by_date={}
    for r in rows:
        by_date[r['attendance_date']]=r

    days=[]
    leading=first_day.weekday()
    for _ in range(leading):
        days.append(None)
    for day_no in range(1,last_day.day+1):
        d=date(year,mon,day_no)
        days.append({'date':d,'record':by_date.get(d)})
    while len(days)%7:
        days.append(None)

    summary={
        'present':sum(1 for r in rows if r.get('status')=='Present'),
        'late':sum(1 for r in rows if r.get('status')=='Late'),
        'half_day':sum(1 for r in rows if r.get('status')=='Half Day'),
        'absent':sum(1 for r in rows if r.get('status') in ('Absent','Leave')),
        'holiday':sum(1 for r in rows if r.get('status')=='Holiday')
    }
    return render_template('admin_attendance_calendar.html',
                           month=month,days=days,employees=employees,
                           selected_employee_id=employee_id,summary=summary)

@app.route('/employee/attendance-calendar')
@employee_required
def employee_attendance_calendar():
    month=(request.args.get('month') or date.today().strftime('%Y-%m'))[:7]
    try:
        year,mon=map(int,month.split('-'))
        first_day=date(year,mon,1)
        last_day=date(year,mon,calendar.monthrange(year,mon)[1])
    except Exception:
        today=date.today()
        year,mon=today.year,today.month
        month=today.strftime('%Y-%m')
        first_day=date(year,mon,1)
        last_day=date(year,mon,calendar.monthrange(year,mon)[1])

    eid=session['employee_id']
    rows=query("""SELECT attendance_date,check_in,check_out,status,working_minutes,
                         late_minutes,overtime_minutes,remarks
                  FROM attendance
                  WHERE employee_id=%s AND attendance_date BETWEEN %s AND %s
                  ORDER BY attendance_date""",(eid,first_day,last_day))
    by_date={r['attendance_date']:r for r in rows}
    days=[None]*first_day.weekday()
    for day_no in range(1,last_day.day+1):
        d=date(year,mon,day_no)
        days.append({'date':d,'record':by_date.get(d)})
    while len(days)%7:
        days.append(None)

    summary={
        'present':sum(1 for r in rows if r.get('status')=='Present'),
        'late':sum(1 for r in rows if r.get('status')=='Late'),
        'half_day':sum(1 for r in rows if r.get('status')=='Half Day'),
        'absent':sum(1 for r in rows if r.get('status') in ('Absent','Leave')),
        'holiday':sum(1 for r in rows if r.get('status')=='Holiday')
    }
    return render_template('employee_attendance_calendar.html',
                           month=month,days=days,summary=summary)

@app.route('/admin/analytics')
@admin_required
def admin_analytics():
    month=(request.args.get('month') or date.today().strftime('%Y-%m'))[:7]
    try:
        year,mon=map(int,month.split('-'))
        month_start=date(year,mon,1)
        month_end=date(year,mon,calendar.monthrange(year,mon)[1])
    except Exception:
        today=date.today()
        year,mon=today.year,today.month
        month=today.strftime('%Y-%m')
        month_start=date(year,mon,1)
        month_end=date(year,mon,calendar.monthrange(year,mon)[1])

    status_rows=query("""SELECT status,COUNT(*) total
                         FROM attendance
                         WHERE attendance_date BETWEEN %s AND %s
                         GROUP BY status ORDER BY total DESC""",(month_start,month_end))
    dept_rows=query("""SELECT COALESCE(d.department_name,'Unassigned') department_name,
                              COUNT(a.id) total,
                              SUM(a.status='Present') present,
                              SUM(a.status='Late') late,
                              SUM(a.status='Half Day') half_day,
                              SUM(a.status IN ('Absent','Leave')) absent
                       FROM attendance a
                       JOIN employees e ON e.id=a.employee_id
                       LEFT JOIN departments d ON d.id=e.department_id
                       WHERE a.attendance_date BETWEEN %s AND %s
                       GROUP BY COALESCE(d.department_name,'Unassigned')
                       ORDER BY total DESC, department_name""",(month_start,month_end))
    totals=query("""SELECT
                       COUNT(*) total,
                       SUM(status='Present') present,
                       SUM(status='Late') late,
                       SUM(status='Half Day') half_day,
                       SUM(status IN ('Absent','Leave')) absent,
                       COALESCE(SUM(working_minutes),0) working_minutes,
                       COALESCE(SUM(overtime_minutes),0) overtime_minutes
                    FROM attendance
                    WHERE attendance_date BETWEEN %s AND %s""",(month_start,month_end),one=True)
    return render_template('admin_analytics.html',month=month,status_rows=status_rows,
                           dept_rows=dept_rows,totals=totals or {})


@app.route('/admin/employees',methods=['GET','POST'])
@admin_required
def admin_employees():
    if request.method=='POST':
        f=request.form
        try:
            login=unique_login_id(f.get('first_name'),f.get('mother_name'),f.get('joining_date'))
            eid=query("""INSERT INTO employees(employee_id,login_id,first_name,last_name,father_name,mother_name,father_phone,mother_phone,dob,gender,email,address,city,district,state,pincode,phone,alternate_phone,blood_group,aadhar,pan,department_id,designation,joining_date,salary,target,status,password) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,NULLIF(%s,''),%s,NULLIF(%s,''),%s,%s,%s,%s,%s,%s,%s,%s,NULLIF(%s,''),NULLIF(%s,''),NULLIF(%s,''),%s,NULLIF(%s,''),NULLIF(%s,''),NULLIF(%s,''),%s,%s)""",(f.get('employee_id') or login,login,f.get('first_name'),f.get('last_name'),f.get('father_name'),f.get('mother_name'),f.get('father_phone'),f.get('mother_phone'),f.get('dob',''),f.get('gender') or None,f.get('email',''),f.get('address'),clean_location(f,'city'),clean_location(f,'district'),f.get('state'),f.get('pincode'),f.get('phone'),f.get('alternate_phone'),f.get('blood_group') or None,f.get('aadhar',''),f.get('pan',''),f.get('department_id',''),f.get('designation'),f.get('joining_date',''),f.get('salary',''),f.get('target',''),f.get('status','Active'),f.get('password','')),commit=True)
            photo=request.files.get('photo')
            if photo and photo.filename:
                ext=photo.filename.rsplit('.',1)[-1].lower() if '.' in photo.filename else ''
                if ext not in ALLOWED: raise ValueError('Photo must be PNG, JPG, JPEG or WEBP.')
                name=f"emp_{eid}_{int(datetime.now().timestamp())}.{ext}";photo.save(os.path.join(app.config['UPLOAD_FOLDER'],name));query("UPDATE employees SET photo=%s WHERE id=%s",(name,eid),commit=True)
            query("UPDATE employees SET employment_stage=%s WHERE id=%s",(f.get('employment_stage') or 'Active',eid),commit=True)
            audit_log('CREATE_EMPLOYEE','Employee',eid,{'login_id':login},eid)
            flash(f'Employee added successfully. Login ID: {login}','success')
        except Exception as e:flash(f'Unable to add employee: {e}','danger')
        return redirect(url_for('admin_employees'))
    emps=query("""SELECT e.*,d.department_name FROM employees e LEFT JOIN departments d ON d.id=e.department_id ORDER BY e.created_at DESC""");deps=query("SELECT * FROM departments WHERE status='Active' ORDER BY department_name")
    return render_template('admin_employees.html',employees=emps,departments=deps)


def employee_profile_payload(form, photo_name=None):
    return {
      'first_name':form.get('first_name'),'last_name':form.get('last_name'),'father_name':form.get('father_name'),'mother_name':form.get('mother_name'),
      'father_phone':form.get('father_phone'),'mother_phone':form.get('mother_phone'),'dob':form.get('dob',''),'gender':form.get('gender') or None,
      'email':form.get('email',''),'address':form.get('address'),'city':clean_location(form,'city'),'district':clean_location(form,'district'),'state':form.get('state'),
      'pincode':form.get('pincode'),'phone':form.get('phone'),'alternate_phone':form.get('alternate_phone'),'blood_group':form.get('blood_group') or None,
      'aadhar':form.get('aadhar',''),'pan':form.get('pan',''),'department_id':form.get('department_id',''),'designation':form.get('designation'),
      'joining_date':form.get('joining_date',''),'salary':form.get('salary',''),'target':form.get('target',''),'status':form.get('status','Active'),
      'password':form.get('password',''),'employment_stage':form.get('employment_stage') or 'Active','inactive_from':(form.get('inactive_from') or '').strip(),
      'photo':photo_name
    }

def apply_employee_profile_payload(eid,payload):
    if payload.get('status')=='Inactive' and not payload.get('inactive_from'):
        raise ValueError('Inactive / Termination Effective Date is required.')
    query("""UPDATE employees SET first_name=%s,last_name=%s,father_name=%s,mother_name=%s,father_phone=%s,mother_phone=%s,dob=NULLIF(%s,''),gender=%s,email=NULLIF(%s,''),address=%s,city=%s,district=%s,state=%s,pincode=%s,phone=%s,alternate_phone=%s,blood_group=%s,aadhar=NULLIF(%s,''),pan=NULLIF(%s,''),department_id=NULLIF(%s,''),designation=%s,joining_date=NULLIF(%s,''),salary=NULLIF(%s,''),target=NULLIF(%s,''),status=%s,password=%s,employment_stage=%s WHERE id=%s""",
          (payload.get('first_name'),payload.get('last_name'),payload.get('father_name'),payload.get('mother_name'),payload.get('father_phone'),payload.get('mother_phone'),payload.get('dob',''),payload.get('gender'),payload.get('email',''),payload.get('address'),payload.get('city'),payload.get('district'),payload.get('state'),payload.get('pincode'),payload.get('phone'),payload.get('alternate_phone'),payload.get('blood_group'),payload.get('aadhar',''),payload.get('pan',''),payload.get('department_id',''),payload.get('designation'),payload.get('joining_date',''),payload.get('salary',''),payload.get('target',''),payload.get('status','Active'),payload.get('password',''),payload.get('employment_stage') or 'Active',eid),commit=True)
    if payload.get('photo'):
        query("UPDATE employees SET photo=%s WHERE id=%s",(payload['photo'],eid),commit=True)
    if payload.get('status')=='Inactive':
        cutoff=payload.get('inactive_from')
        query("UPDATE employees SET inactive_from=%s WHERE id=%s",(cutoff,eid),commit=True)
        query("""DELETE FROM attendance WHERE employee_id=%s AND attendance_date >= %s AND (remarks LIKE 'Auto Sunday%%' OR remarks LIKE 'Auto Leave:%%' OR remarks LIKE 'Auto Absent%%' OR remarks LIKE 'Auto Present%%')""",(eid,cutoff),commit=True)
    else:
        query("UPDATE employees SET inactive_from=NULL WHERE id=%s",(eid,),commit=True)

def create_approval(request_type,entity_type,entity_id,title,payload,stage='Manager',hr_approved=True):
    return query("""INSERT INTO approval_requests(request_type,entity_type,entity_id,title,payload_json,status,current_stage,submitted_by,hr_by,hr_at)
                    VALUES(%s,%s,%s,%s,%s,'Pending',%s,%s,%s,%s)""",
                 (request_type,entity_type,str(entity_id),title,json.dumps(payload,ensure_ascii=False,default=str),stage,session['admin_id'],session['admin_id'] if hr_approved else None,datetime.now() if hr_approved else None),commit=True)

@app.route('/admin/employees/<int:eid>',methods=['GET','POST'])
@admin_required
def employee_edit(eid):
    emp=query("SELECT * FROM employees WHERE id=%s",(eid,),one=True)
    if not emp:return ('Not found',404)
    if request.method=='POST':
        role=current_admin_role()
        if role not in ('HR','Super Admin'):
            flash('Employee profile changes must be initiated by HR. Manager and Owner review them in Approval Center.','warning')
            return redirect(url_for('employee_edit',eid=eid))
        try:
            photo_name=None
            photo=request.files.get('photo')
            if photo and photo.filename:
                ext=photo.filename.rsplit('.',1)[-1].lower() if '.' in photo.filename else ''
                if ext not in ALLOWED: raise ValueError('Photo must be PNG, JPG, JPEG or WEBP.')
                photo_name=f"emp_{eid}_{int(datetime.now().timestamp())}.{ext}"
                photo.save(os.path.join(app.config['UPLOAD_FOLDER'],photo_name))
            payload=employee_profile_payload(request.form,photo_name)
            if role=='Super Admin':
                apply_employee_profile_payload(eid,payload)
                audit_log('UPDATE_EMPLOYEE','Employee',eid,{'mode':'Super Admin direct update'},eid)
                flash('Employee profile updated.','success')
            else:
                existing=query("SELECT id FROM approval_requests WHERE request_type='Employee Profile Update' AND entity_id=%s AND status='Pending'",(str(eid),),one=True)
                if existing: raise ValueError('A profile update for this employee is already awaiting approval.')
                create_approval('Employee Profile Update','Employee',eid,f"Profile update · {emp.get('first_name')} {emp.get('last_name') or ''}",payload,'Manager',True)
                audit_log('SUBMIT_EMPLOYEE_UPDATE','Employee',eid,{'approval_required':True},eid)
                flash('Profile update submitted. Manager approval is now required, followed by Owner final approval.','success')
            return redirect(url_for('employee_edit',eid=eid))
        except Exception as e:
            flash(f'Unable to save employee profile: {e}','danger')
            return redirect(url_for('employee_edit',eid=eid))
    deps=query("SELECT * FROM departments WHERE status='Active' ORDER BY department_name")
    docs=query("SELECT * FROM employee_documents WHERE employee_id=%s ORDER BY uploaded_at DESC",(eid,))
    pending=query("SELECT * FROM approval_requests WHERE request_type='Employee Profile Update' AND entity_id=%s AND status='Pending' ORDER BY created_at DESC LIMIT 1",(str(eid),),one=True)
    return render_template('employee_profile_admin.html',employee=emp,departments=deps,documents=docs,digilocker_configured=dl.configured(),pending_profile_approval=pending)

@app.route('/admin/employees/<int:eid>/view')
@admin_required
def employee_view(eid):
    emp=query("""SELECT e.*,d.department_name FROM employees e LEFT JOIN departments d ON d.id=e.department_id WHERE e.id=%s""",(eid,),one=True)
    if not emp:return ('Not found',404)
    docs=query("SELECT * FROM employee_documents WHERE employee_id=%s ORDER BY uploaded_at DESC",(eid,));return render_template('employee_view.html',employee=emp,documents=docs)

@app.post('/admin/employees/<int:eid>/documents')
@admin_required
def employee_document_upload(eid):
    f=request.files.get('document');title=request.form.get('title','').strip();dtype=request.form.get('document_type','Other')
    if not f or not f.filename:flash('Choose a document file.','warning');return redirect(url_for('employee_edit',eid=eid)+'#documents')
    ext=f.filename.rsplit('.',1)[-1].lower() if '.' in f.filename else ''
    if ext not in DOC_ALLOWED:flash('Unsupported document type.','danger');return redirect(url_for('employee_edit',eid=eid)+'#documents')
    safe_original=secure_filename(f.filename);title=(title or f.filename)[:150];dtype=(dtype or 'Other')[:80];name=f"doc_{eid}_{int(datetime.now().timestamp())}_{safe_original}"[:255];path=os.path.join(DOCUMENT_UPLOAD,name);f.save(path)
    try:query("INSERT INTO employee_documents(employee_id,title,document_type,file_name,original_name,verification_status,verification_source) VALUES(%s,%s,%s,%s,%s,'Pending','Manual Upload')",(eid,title,dtype,name,f.filename[:255]),commit=True)
    except Exception:
        try:os.remove(path)
        except OSError:pass
        raise
    flash('Document uploaded. DigiLocker verification will remain pending until official API credentials are connected.','success');return redirect(url_for('employee_edit',eid=eid)+'#documents')


@app.get('/admin/employees/<int:eid>/digilocker/start')
@admin_required
def digilocker_start(eid):
    if not dl.configured():
        flash('DigiLocker Requester credentials are not configured yet. Complete API Setu partner onboarding first.','warning')
        return redirect(url_for('employee_edit',eid=eid)+'#documents')

    emp=query("SELECT id,first_name,last_name FROM employees WHERE id=%s",(eid,),one=True)
    if not emp:
        return ('Employee not found',404)

    state=dl.make_state()
    verifier,challenge=dl.make_pkce()
    expires=datetime.now()+timedelta(minutes=15)

    # Remove stale/incomplete flows for this admin+employee.
    query("DELETE FROM digilocker_oauth_sessions WHERE expires_at<NOW() OR (employee_id=%s AND admin_id=%s)",
          (eid,session['admin_id']),commit=True)
    query("""INSERT INTO digilocker_oauth_sessions
             (state,employee_id,admin_id,code_verifier,expires_at)
             VALUES(%s,%s,%s,%s,%s)""",
          (state,eid,session['admin_id'],verifier,expires),commit=True)
    return redirect(dl.authorization_url(state,challenge))

@app.get('/admin/digilocker/callback')
@admin_required
def digilocker_callback():
    error=request.args.get('error')
    state=(request.args.get('state') or '').strip()
    if error:
        flash('DigiLocker authorization was not completed: '+(request.args.get('error_description') or error),'warning')
        return redirect(url_for('admin_employees'))

    flow=query("""SELECT * FROM digilocker_oauth_sessions
                  WHERE state=%s AND admin_id=%s AND expires_at>NOW()""",
               (state,session['admin_id']),one=True)
    if not flow:
        flash('DigiLocker verification session is invalid or expired. Please start again.','danger')
        return redirect(url_for('admin_employees'))

    code=(request.args.get('code') or '').strip()
    if not code:
        flash('DigiLocker did not return an authorization code.','danger')
        return redirect(url_for('employee_edit',eid=flow['employee_id'])+'#documents')

    try:
        token=dl.exchange_code(code,flow['code_verifier'])
        access=token.get('access_token')
        if not access:
            raise ValueError('DigiLocker access token was not returned.')
        docs=dl.list_issued_documents(access)
        access_enc=dl.encrypt_token(access,app.config['SECRET_KEY'])
        refresh_enc=dl.encrypt_token(token.get('refresh_token'),app.config['SECRET_KEY'])
        query("""UPDATE digilocker_oauth_sessions
                 SET access_token_enc=%s,refresh_token_enc=%s,issued_docs_json=%s,
                     expires_at=DATE_ADD(NOW(),INTERVAL 10 MINUTE)
                 WHERE id=%s""",
              (access_enc,refresh_enc,json.dumps(docs,ensure_ascii=False),flow['id']),commit=True)
        return redirect(url_for('digilocker_select',flow_id=flow['id']))
    except Exception as e:
        flash(f'DigiLocker connection failed: {e}','danger')
        return redirect(url_for('employee_edit',eid=flow['employee_id'])+'#documents')

@app.get('/admin/digilocker/<int:flow_id>/documents')
@admin_required
def digilocker_select(flow_id):
    flow=query("""SELECT d.*,e.first_name,e.last_name
                  FROM digilocker_oauth_sessions d
                  JOIN employees e ON e.id=d.employee_id
                  WHERE d.id=%s AND d.admin_id=%s AND d.expires_at>NOW()""",
               (flow_id,session['admin_id']),one=True)
    if not flow:
        flash('DigiLocker verification session expired.','warning')
        return redirect(url_for('admin_employees'))
    try: docs=json.loads(flow.get('issued_docs_json') or '[]')
    except Exception: docs=[]
    return render_template('digilocker_select.html',flow=flow,documents=docs)

@app.post('/admin/digilocker/<int:flow_id>/import')
@admin_required
def digilocker_import(flow_id):
    flow=query("""SELECT * FROM digilocker_oauth_sessions
                  WHERE id=%s AND admin_id=%s AND expires_at>NOW()""",
               (flow_id,session['admin_id']),one=True)
    if not flow:
        flash('DigiLocker verification session expired.','warning')
        return redirect(url_for('admin_employees'))

    try:
        docs=json.loads(flow.get('issued_docs_json') or '[]')
        index=int(request.form.get('doc_index','-1'))
        if index<0 or index>=len(docs):
            raise ValueError('Invalid DigiLocker document selection.')
        meta=docs[index] or {}
        uri=(meta.get('uri') or '').strip()
        if not uri:
            raise ValueError('Selected DigiLocker document has no URI.')

        access=dl.decrypt_token(flow['access_token_enc'],app.config['SECRET_KEY'])
        content,ctype=dl.fetch_document(access,uri)
        ext=dl.extension_for(ctype)
        name=f"digilocker_{flow['employee_id']}_{int(datetime.now().timestamp())}_{secrets.token_hex(4)}{ext}"
        path=os.path.join(DOCUMENT_UPLOAD,name)
        with open(path,'wb') as handle:
            handle.write(content)

        title=(meta.get('name') or meta.get('description') or 'DigiLocker Verified Document')[:150]
        dtype=(meta.get('doctype') or meta.get('description') or meta.get('name') or 'DigiLocker Document')[:80]
        original=(meta.get('name') or os.path.basename(uri) or name)[:255]

        existing=query("""SELECT id FROM employee_documents
                          WHERE employee_id=%s AND verification_source='DigiLocker'
                          AND external_document_id=%s LIMIT 1""",
                       (flow['employee_id'],uri),one=True)
        if existing:
            try: os.remove(path)
            except OSError: pass
            flash('This DigiLocker document is already saved for the employee.','info')
        else:
            query("""INSERT INTO employee_documents
                     (employee_id,title,document_type,file_name,original_name,
                      verification_status,verification_source,external_document_id,verified_at)
                     VALUES(%s,%s,%s,%s,%s,'Verified','DigiLocker',%s,NOW())""",
                  (flow['employee_id'],title,dtype,name,original,uri),commit=True)
            flash('Verified DigiLocker document saved successfully.','success')
        return redirect(url_for('digilocker_select',flow_id=flow_id))
    except Exception as e:
        flash(f'Unable to save DigiLocker document: {e}','danger')
        return redirect(url_for('digilocker_select',flow_id=flow_id))

@app.post('/admin/digilocker/<int:flow_id>/finish')
@admin_required
def digilocker_finish(flow_id):
    flow=query("""SELECT * FROM digilocker_oauth_sessions
                  WHERE id=%s AND admin_id=%s""",(flow_id,session['admin_id']),one=True)
    eid=flow['employee_id'] if flow else None
    if flow:
        try:
            access=dl.decrypt_token(flow.get('access_token_enc'),app.config['SECRET_KEY'])
            dl.revoke(access)
        except Exception:
            pass
        query("DELETE FROM digilocker_oauth_sessions WHERE id=%s",(flow_id,),commit=True)
    return redirect(url_for('employee_edit',eid=eid)+'#documents') if eid else redirect(url_for('admin_employees'))

@app.post('/admin/documents/<int:did>/delete')
@admin_required
def employee_document_delete(did):
    d=query("SELECT * FROM employee_documents WHERE id=%s",(did,),one=True)
    if d:
        try:os.remove(os.path.join(DOCUMENT_UPLOAD,d['file_name']))
        except OSError:pass
        query("DELETE FROM employee_documents WHERE id=%s",(did,),commit=True);flash('Document deleted.','success');return redirect(url_for('employee_edit',eid=d['employee_id'])+'#documents')
    flash('Document not found.','warning');return redirect(url_for('admin_employees'))

@app.post('/admin/employees/<int:eid>/delete')
@admin_required
def employee_delete(eid):
    emp=query("SELECT id,status FROM employees WHERE id=%s",(eid,),one=True)
    if not emp:
        flash('Employee not found.','warning')
        return redirect(url_for('admin_employees'))
    query("UPDATE employees SET status='Inactive',employment_stage='Archived',archived_at=NOW(),archived_by=%s WHERE id=%s",
          (session['admin_id'],eid),commit=True)
    query("INSERT INTO employee_exit_checklist(employee_id,updated_by) VALUES(%s,%s) ON DUPLICATE KEY UPDATE updated_by=VALUES(updated_by)",
          (eid,session['admin_id']),commit=True)
    audit_log('ARCHIVE_EMPLOYEE','Employee',eid,{'previous_status':emp.get('status')},eid)
    flash('Employee archived. Historical attendance, documents and payroll records were preserved.','success')
    return redirect(url_for('admin_employees'))


def sync_sunday_salary_days(month_start, month_end):
    """
    Company Sunday salary rule:
      - Sunday is automatically credited as Present.
      - If BOTH adjacent Saturday and Monday are Absent, Sunday becomes Absent.
      - Inactive / terminated employees are excluded from their effective date onward.
      - Existing real/manual Sunday attendance is never overwritten.
      - Auto-created Sunday rows carry a dedicated remark and are recalculated safely.
    """

    auto_remark = 'Auto Present'
    d = month_start

    while d <= month_end:

        if d.weekday() != 6:   # Sunday
            d += timedelta(days=1)
            continue

        sat = d - timedelta(days=1)
        mon = d + timedelta(days=1)

        # Only employees who were still active on this Sunday.
        employees = query("""
            SELECT id, joining_date, status, inactive_from
            FROM employees
            WHERE
                joining_date IS NULL
                OR joining_date <= %s
        """, (d,))

        for emp in employees:

            joining = emp.get('joining_date')
            inactive_from = emp.get('inactive_from')

            if joining and joining > d:
                continue

            # Employee is no longer eligible from inactive_from date onward.
            if inactive_from and d >= inactive_from:
                # Clean previously auto-created Sunday row if it exists.
                query("""
                    DELETE FROM attendance
                    WHERE employee_id=%s
                      AND attendance_date=%s
                      AND remarks=%s
                """, (emp['id'], d, auto_remark), commit=True)

                continue

            # If status is already Inactive but inactive_from is missing,
            # safest behavior is to exclude the employee completely.
            if emp.get('status') != 'Active' and not inactive_from:
                continue

            eid = emp['id']

            neighbors = query("""
                SELECT attendance_date, status
                FROM attendance
                WHERE employee_id=%s
                  AND attendance_date IN (%s,%s)
            """, (eid, sat, mon))

            statuses = {
                r['attendance_date']:
                    ('Absent' if r['status'] == 'Leave' else r['status'])
                for r in neighbors
            }

            derived = (
                'Absent'
                if statuses.get(sat) == 'Absent'
                and statuses.get(mon) == 'Absent'
                else 'Present'
            )

            existing = query("""
                SELECT id,status,check_in,check_out,remarks
                FROM attendance
                WHERE employee_id=%s
                  AND attendance_date=%s
            """, (eid, d), one=True)

            if existing:

                # Only recalculate rows previously auto-generated by this rule,
                # or legacy blank Holiday/Leave Sundays.
                # Never replace real punch/manual attendance.
                is_auto = (existing.get('remarks') or '') == auto_remark

                legacy_sunday = (
                    existing.get('status') in ('Holiday', 'Leave')
                    and not existing.get('check_in')
                    and not existing.get('check_out')
                )

                if is_auto or legacy_sunday:
                    query("""
                        UPDATE attendance
                        SET status=%s,
                            check_in=NULL,
                            check_out=NULL,
                            working_minutes=0,
                            late_minutes=0,
                            overtime_minutes=0,
                            early_exit_minutes=0,
                            remarks=%s
                        WHERE id=%s
                    """, (derived, auto_remark, existing['id']), commit=True)

            else:

                query("""
                    INSERT INTO attendance
                    (
                        employee_id,
                        attendance_date,
                        status,
                        working_minutes,
                        late_minutes,
                        overtime_minutes,
                        early_exit_minutes,
                        remarks
                    )
                    VALUES(%s,%s,%s,0,0,0,0,%s)
                """, (eid, d, derived, auto_remark), commit=True)

        d += timedelta(days=1)

def apply_approved_leave_as_absent(leave_id):
    """
    Approved leave is represented in Attendance as Absent.
    Half-day leave remains Half Day.
    Sunday itself is governed by the Sunday salary/sandwich rule.
    """
    lr=query("""SELECT employee_id,start_date,end_date,day_type,status
                FROM leave_requests WHERE id=%s""",(leave_id,),one=True)
    if not lr or lr.get('status')!='Approved':
        return

    current=lr['start_date']
    end=lr['end_date']
    is_half=str(lr.get('day_type') or '') in ('First Half','Second Half')
    while current<=end:
        if current.weekday()!=6:  # Sunday is calculated from Sat + Mon instead.
            desired='Half Day' if is_half else 'Absent'
            existing=query("""SELECT id,check_in,check_out,status,remarks
                              FROM attendance
                              WHERE employee_id=%s AND attendance_date=%s""",
                           (lr['employee_id'],current),one=True)
            if not existing:
                query("""INSERT INTO attendance
                         (employee_id,attendance_date,status,working_minutes,late_minutes,
                          overtime_minutes,early_exit_minutes,remarks)
                         VALUES(%s,%s,%s,0,0,0,0,%s)""",
                      (lr['employee_id'],current,desired,'Approved leave - attendance counted as Absent'),commit=True)
            elif not existing.get('check_in') and not existing.get('check_out') and existing.get('status') in ('Leave','Absent','Holiday'):
                query("""UPDATE attendance SET status=%s,remarks=%s WHERE id=%s""",
                      (desired,'Approved leave - attendance counted as Absent',existing['id']),commit=True)
        current += timedelta(days=1)

@app.route('/admin/attendance')
@admin_required
def admin_attendance():
    month=request.args.get('month') or date.today().strftime('%Y-%m')
    try: sync_company_attendance_rules()
    except Exception: pass
    try:
        _y,_m=map(int,month.split('-'))
        _start=date(_y,_m,1)
        _end=date(_y,_m,calendar.monthrange(_y,_m)[1])
        sync_sunday_salary_days(_start,_end)
    except Exception:
        pass
    employee_id=(request.args.get('employee_id') or '').strip()
    exception=(request.args.get('exception') or '').strip()
    params=[month]
    where="DATE_FORMAT(a.attendance_date,'%Y-%m')=%s"
    if employee_id.isdigit():
        where += " AND a.employee_id=%s"
        params.append(int(employee_id))
    if exception=='Late': where += " AND a.status='Late'"
    elif exception=='Half Day': where += " AND a.status='Half Day'"
    elif exception=='Absent': where += " AND a.status IN ('Absent','Leave')"
    elif exception=='Missing Checkout': where += " AND a.check_in IS NOT NULL AND a.check_out IS NULL AND a.attendance_date<CURDATE()"
    elif exception=='Manual Edited': where += " AND EXISTS (SELECT 1 FROM attendance_audit aa WHERE aa.attendance_id=a.id)"
    rows=query(f"""SELECT a.*,CONCAT(e.first_name,' ',COALESCE(e.last_name,'')) employee_name,e.login_id,e.designation FROM attendance a JOIN employees e ON e.id=a.employee_id WHERE {where} ORDER BY a.attendance_date DESC,a.check_in DESC""",tuple(params))
    employees=query("SELECT id,login_id,first_name,last_name FROM employees WHERE status='Active' ORDER BY first_name,last_name")
    return render_template('admin_attendance.html',rows=rows,month=month,employees=employees,selected_employee_id=employee_id,exception_filter=exception,month_locked=month_is_locked(month),anomaly_count=attendance_anomaly_count(month))

@app.post('/admin/attendance/save')
@admin_required
def admin_attendance_save():
    f=request.form
    try:
        target_month=(f.get('attendance_date') or '')[:7]
        if target_month and month_is_locked(target_month):
            raise ValueError('This attendance month is closed. Reopen it before making corrections.')
        aid=int(f.get('attendance_id') or 0);eid=int(f.get('employee_id'));att_date=datetime.strptime(f.get('attendance_date'),'%Y-%m-%d').date();check_in=parse_dt(att_date,f.get('check_in'));check_out=parse_dt(att_date,f.get('check_out'))
        if check_in and check_out and check_out<check_in:raise ValueError('Check-out cannot be before check-in.')
        working,late,overtime,early=attendance_metrics(att_date,check_in,check_out);status=f.get('status') or 'Present';status='Absent' if status=='Leave' else status;reason=(f.get('correction_reason') or '').strip();remarks=(f.get('remarks') or '').strip()
        if not reason:raise ValueError('Correction reason is required for audit history.')
        old=query("SELECT * FROM attendance WHERE id=%s",(aid,),one=True) if aid else None
        if aid:
            query("UPDATE attendance SET employee_id=%s,attendance_date=%s,check_in=%s,check_out=%s,working_minutes=%s,late_minutes=%s,overtime_minutes=%s,early_exit_minutes=%s,status=%s,remarks=%s WHERE id=%s",(eid,att_date,check_in,check_out,working,late,overtime,early,status,remarks,aid),commit=True);record_id=aid;action='EDIT'
        else:
            existing=query("SELECT id FROM attendance WHERE employee_id=%s AND attendance_date=%s",(eid,att_date),one=True)
            if existing:raise ValueError('Attendance already exists for this employee/date. Use Edit instead.')
            record_id=query("INSERT INTO attendance(employee_id,attendance_date,check_in,check_out,working_minutes,late_minutes,overtime_minutes,early_exit_minutes,status,remarks) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",(eid,att_date,check_in,check_out,working,late,overtime,early,status,remarks),commit=True);action='ADD'
        new=query("SELECT * FROM attendance WHERE id=%s",(record_id,),one=True)
        query("INSERT INTO attendance_audit(attendance_id,employee_id,admin_id,action_type,reason,old_data,new_data) VALUES(%s,%s,%s,%s,%s,%s,%s)",(record_id,eid,session['admin_id'],action,reason,json.dumps(old,default=str) if old else None,json.dumps(new,default=str)),commit=True)
        audit_log('ATTENDANCE_'+action,'Attendance',record_id,{'reason':reason,'date':str(att_date)},eid)
        notify_employee(eid,'Attendance updated',f'Your attendance for {att_date.strftime("%d %b %Y")} was updated by an administrator.','Attendance')
        flash('Attendance corrected successfully and recorded in audit history.','success')
    except Exception as e:flash(f'Unable to save attendance correction: {e}','danger')
    return redirect(url_for('admin_attendance',month=f.get('attendance_date','')[:7] or date.today().strftime('%Y-%m')))

@app.post('/admin/attendance/import')
@admin_required
def admin_attendance_import():
    upload=request.files.get('attendance_file')
    if not upload or not upload.filename.lower().endswith('.xlsx'):
        flash('Please choose the July attendance .xlsx file.','warning');return redirect(url_for('admin_attendance'))
    temp=os.path.join(app.config['UPLOAD_FOLDER'],f"attendance_import_{secrets.token_hex(6)}.xlsx");upload.save(temp)
    inserted=skipped=0;unmatched=set()
    try:
        wb=load_workbook(temp,data_only=True);ws=wb.active;s=settings();report_time=as_time(s.get('last_reporting'),time(9,45));half_time=as_time(s.get('half_day_after'),time(10,30))
        employees=query("SELECT id,first_name,last_name FROM employees")
        empmap={re.sub(r'\s+',' ',(e['first_name']+' '+(e.get('last_name') or '')).strip()).casefold():e['id'] for e in employees}
        headers=[]
        for c in range(4,ws.max_column+1):
            v=ws.cell(1,c).value
            if isinstance(v,datetime):headers.append((c,v.date()))
            elif isinstance(v,date):headers.append((c,v))
            elif isinstance(v,(int,float)):
                try:headers.append((c,from_excel(v).date()))
                except Exception:pass
        time_rx=re.compile(r'(\d{1,2}:\d{2}(?::\d{2})?)\s*(AM|PM)',re.I)
        for r in range(2,ws.max_row+1):
            name=str(ws.cell(r,3).value or '').strip();key=re.sub(r'\s+',' ',name).casefold();eid=empmap.get(key)
            if not eid:
                if name:unmatched.add(name)
                continue
            for c,att_date in headers:
                raw=ws.cell(r,c).value
                if raw is None:continue
                status=None;check_in=None;remarks='';late_reason=None
                if isinstance(raw,datetime):check_in=datetime.combine(att_date,raw.time())
                elif isinstance(raw,time):check_in=datetime.combine(att_date,raw)
                elif isinstance(raw,(int,float)) and 0<=raw<1:
                    seconds=round(raw*86400)%86400;check_in=datetime.combine(att_date,time(seconds//3600,(seconds%3600)//60,seconds%60))
                else:
                    text=str(raw).strip()
                    if not text or text=='-':continue
                    low=text.casefold()
                    if low in {'ab','absent'}:status='Absent'
                    elif 'on leave' in low:status='Absent'
                    elif 'office holiday' in low.replace('\n',' ') or low=='sunday':status='Holiday'
                    elif low in {'terminate','login missing'}:continue
                    else:
                        m=time_rx.search(text)
                        if m:
                            fmt='%I:%M:%S %p' if m.group(1).count(':')==2 else '%I:%M %p';tm=datetime.strptime(m.group(1)+' '+m.group(2).upper(),fmt).time();check_in=datetime.combine(att_date,tm);lines=[x.strip() for x in text.splitlines()[1:] if x.strip() and x.strip().casefold()!='(late)'];remarks=' | '.join(lines)[:255];late_reason=remarks[:500] if remarks else None
                        else:continue
                late=0
                if check_in:
                    report=datetime.combine(att_date,report_time);late=max(0,int((check_in-report).total_seconds()//60));status='Half Day' if check_in.time()>half_time else ('Late' if late else 'Present')
                sql="INSERT IGNORE INTO attendance(employee_id,attendance_date,check_in,late_minutes,status,remarks,late_reason) VALUES(%s,%s,%s,%s,%s,%s,%s)"
                new_id=query(sql,(eid,att_date,check_in,late,status or 'Present',remarks or None,late_reason),commit=True)
                if new_id:inserted+=1
                else:skipped+=1
        msg=f'Attendance import complete: {inserted} new records added, {skipped} existing records preserved.'
        if unmatched:msg+=f" Unmatched employees: {', '.join(sorted(unmatched))}."
        flash(msg,'success' if inserted else 'warning')
    except Exception as e:flash(f'Attendance import failed: {e}','danger')
    finally:
        try:os.remove(temp)
        except OSError:pass
    return redirect(url_for('admin_attendance',month='2026-07'))

@app.route('/admin/departments',methods=['GET','POST'])
@admin_required
def admin_departments():
    if request.method=='POST':
        f=request.form
        if f.get('id'):query("UPDATE departments SET department_name=%s,description=%s,status=%s WHERE id=%s",(f.get('department_name'),f.get('description'),f.get('status'),f.get('id')),commit=True)
        else:query("INSERT INTO departments(department_name,description,status) VALUES(%s,%s,%s)",(f.get('department_name'),f.get('description'),f.get('status','Active')),commit=True)
        flash('Department saved.','success');return redirect(url_for('admin_departments'))
    rows=query("""SELECT d.*,COUNT(e.id) employee_count FROM departments d LEFT JOIN employees e ON e.department_id=d.id GROUP BY d.id ORDER BY d.department_name""");return render_template('admin_departments.html',departments=rows)

@app.post('/admin/departments/<int:did>/delete')
@admin_required
def department_delete(did):
    try:query("UPDATE employees SET department_id=NULL WHERE department_id=%s",(did,),commit=True);query("DELETE FROM departments WHERE id=%s",(did,),commit=True);flash('Department deleted successfully.','success')
    except Exception as e:flash(f'Unable to delete department: {e}','danger')
    return redirect(url_for('admin_departments'))

@app.route('/admin/holidays',methods=['GET','POST'])
@admin_required
def admin_holidays():
    if request.method=='POST':
        f=request.form;hid=f.get('holiday_id')
        try:
            if hid:query("UPDATE holidays SET holiday_date=%s,holiday_name=%s,holiday_type=%s,description=%s,status=%s WHERE id=%s",(f.get('holiday_date'),f.get('holiday_name'),f.get('holiday_type'),f.get('description'),f.get('status','Active'),hid),commit=True)
            else:query("INSERT INTO holidays(holiday_date,holiday_name,holiday_type,description,status) VALUES(%s,%s,%s,%s,%s)",(f.get('holiday_date'),f.get('holiday_name'),f.get('holiday_type'),f.get('description'),f.get('status','Active')),commit=True)
            flash('Holiday saved.','success')
        except Exception as e:flash(f'Unable to save holiday: {e}','danger')
        return redirect(url_for('admin_holidays'))
    rows=query("SELECT * FROM holidays ORDER BY holiday_date DESC");return render_template('admin_holidays.html',holidays=rows)

@app.post('/admin/holidays/<int:hid>/delete')
@admin_required
def holiday_delete(hid):
    try:query("DELETE FROM holidays WHERE id=%s",(hid,),commit=True);flash('Holiday deleted successfully.','success')
    except Exception as e:flash(f'Unable to delete holiday: {e}','danger')
    return redirect(url_for('admin_holidays'))

@app.route('/admin/leaves')
@admin_required
def admin_leaves():
    rows=query("""SELECT lr.*,CONCAT(e.first_name,' ',COALESCE(e.last_name,'')) employee_name,e.login_id,e.designation,lt.name leave_type FROM leave_requests lr JOIN employees e ON e.id=lr.employee_id JOIN leave_types lt ON lt.id=lr.leave_type_id ORDER BY FIELD(lr.status,'Pending','Approved','Rejected','Cancelled'),lr.created_at DESC""");summary=query("SELECT COUNT(*) total,SUM(status='Pending') pending,SUM(status='Approved') approved,SUM(status='Rejected') rejected FROM leave_requests",one=True);return render_template('admin_leaves.html',rows=rows,summary=summary)
@app.post('/admin/leaves/<int:lid>/decision')
@admin_required
def leave_decision(lid):
    status=request.form.get('status'); remark=(request.form.get('admin_remark') or '').strip()
    if status not in {'Approved','Rejected'}:
        flash('Choose Approve or Reject.','warning'); return redirect(url_for('admin_leaves'))
    role=current_admin_role()
    lr_full=query("SELECT employee_id,start_date,end_date FROM leave_requests WHERE id=%s",(lid,),one=True)
    if not lr_full:
        flash('Leave request not found.','danger'); return redirect(url_for('admin_leaves'))
    if role=='Super Admin':
        if status=='Approved':
            try: finalize_leave_approval(lid,remark)
            except Exception as e: flash(f'Leave approved, but attendance synchronization needs review: {e}','warning')
        else:
            query("UPDATE leave_requests SET status='Rejected',admin_remark=%s,reviewed_by=%s,reviewed_at=NOW() WHERE id=%s",(remark,session['admin_id'],lid),commit=True)
            notify_employee(lr_full['employee_id'],'Leave Rejected','Your leave request has been rejected.'+(f' Remark: {remark}' if remark else ''),'Leave')
            audit_log('LEAVE_REJECTED','LeaveRequest',lid,{'remark':remark},lr_full['employee_id'])
        flash(f'Leave request {status.lower()} by Super Admin.','success'); return redirect(url_for('admin_leaves'))
    if role!='HR':
        flash('Leave workflow must be initiated by HR. Manager and Owner act from Approval Center.','warning'); return redirect(url_for('admin_leaves'))
    if status=='Rejected':
        query("UPDATE leave_requests SET status='Rejected',admin_remark=%s,reviewed_by=%s,reviewed_at=NOW() WHERE id=%s",(remark,session['admin_id'],lid),commit=True)
        notify_employee(lr_full['employee_id'],'Leave Rejected','Your leave request has been rejected by HR.'+(f' Remark: {remark}' if remark else ''),'Leave')
        audit_log('LEAVE_REJECTED_HR','LeaveRequest',lid,{'remark':remark},lr_full['employee_id'])
        flash('Leave request rejected by HR.','warning'); return redirect(url_for('admin_leaves'))
    existing=query("SELECT id FROM approval_requests WHERE request_type='Leave Approval' AND entity_id=%s AND status='Pending'",(str(lid),),one=True)
    if existing:
        flash('This leave request is already in the approval workflow.','warning'); return redirect(url_for('admin_leaves'))
    create_approval('Leave Approval','LeaveRequest',lid,f'Leave request #{lid}',{'remark':remark},'Manager',True)
    audit_log('SUBMIT_LEAVE_APPROVAL','LeaveRequest',lid,{'remark':remark},lr_full['employee_id'])
    flash('HR approval recorded. Manager approval is now required, followed by Owner final approval.','success')
    return redirect(url_for('admin_leaves'))


@app.route('/admin/allocation')
@admin_required
def admin_allocation():
    snap=query("SELECT * FROM admin_allocation_snapshots ORDER BY imported_at DESC,id DESC LIMIT 1",one=True)
    tables=[]
    if snap:
        try: tables=json.loads(snap.get('tables_json') or '[]')
        except Exception: tables=[]

    def table_kind(table):
        headers=[str(x or '').strip().upper().replace(' ','') for x in (table.get('headers') or [])]
        joined='|'.join(headers)
        if 'STATE' in headers:
            return 'state'
        if 'CALLER' in headers:
            return 'caller'
        if any(('FRESH' in h and 'STAB' in h) or h in ('FRESH/STAB','FRESHSTAB') for h in headers):
            return 'fresh'
        # Fallback from visible data values if header text is unusual.
        rows=table.get('rows') or []
        preview='|'.join(str(c or '').strip().upper() for row in rows[:4] for c in row[:3])
        if 'FRESH' in preview or 'STAB' in preview:
            return 'fresh'
        return 'unknown'

    title_map={
        'state':'State-wise Performance',
        'caller':'Caller-wise Performance',
        'fresh':'Fresh / Stab Performance',
        'unknown':'Performance'
    }
    order_map={'state':0,'caller':1,'fresh':2,'unknown':99}

    for table in tables:
        table['_kind']=table_kind(table)
    tables=sorted(tables,key=lambda t:order_map.get(t.get('_kind'),99))

    for display_index, table in enumerate(tables, start=1):
        table['display_index']=display_index
        table['display_title']=title_map.get(table.get('_kind'),'Performance')

        rows=table.get('rows') or []
        display=[]
        i=0
        while i < len(rows):
            row=rows[i]
            first=str(row[0] or '').strip() if row else ''
            second=str(row[1] or '').strip() if len(row)>1 else ''
            has_values=any(str(x or '').strip() for x in row[2:]) if len(row)>2 else False

            # Existing manager subtotal rows keep "Total".
            # The final overall row merges first two cells but stays EMPTY, as requested.
            is_total=first.lower() in ('grand total','total') or (not first and not second and has_values)
            if is_total:
                final_row=(i == len(rows)-1)
                display.append({
                    'cells':row,
                    'is_total':True,
                    'is_final_total':final_row,
                    'total_label':'' if final_row else 'Total',
                    'show_ro':False,
                    'rowspan':0
                })
                i+=1
                continue

            if first:
                j=i+1
                while j<len(rows):
                    next_first=str((rows[j][0] if rows[j] else '') or '').strip()
                    next_second=str((rows[j][1] if len(rows[j])>1 else '') or '').strip()
                    next_has_values=any(str(x or '').strip() for x in rows[j][2:]) if len(rows[j])>2 else False
                    next_is_total=next_first.lower() in ('grand total','total') or (not next_first and not next_second and next_has_values)
                    if next_first or next_is_total:
                        break
                    j+=1
                display.append({'cells':row,'is_total':False,'is_final_total':False,'show_ro':True,'rowspan':j-i})
                for k in range(i+1,j):
                    display.append({'cells':rows[k],'is_total':False,'is_final_total':False,'show_ro':False,'rowspan':0})
                i=j
            else:
                display.append({'cells':row,'is_total':False,'is_final_total':False,'show_ro':False,'rowspan':0})
                i+=1

        table['display_rows']=display

    return render_template('admin_allocation.html',snapshot=snap,tables=tables)

@app.post('/admin/allocation/refresh')
@admin_required
def admin_allocation_refresh():
    try:
        from allocation_sync import sync_admin_allocation
        result=sync_admin_allocation(force=True)
        flash('Performance Tracking refreshed from the latest Koofr workbook.','success')
    except Exception as e:
        flash(f'Performance refresh failed: {e}','danger')
    return redirect(url_for('admin_allocation'))


@app.route('/admin/search')
@admin_required
def admin_global_search():
    q=(request.args.get('q') or '').strip()
    employees=[];documents=[]
    if q:
        like=f"%{q}%"
        employees=query("""SELECT e.id,e.login_id,e.first_name,e.last_name,e.phone,e.designation,d.department_name
                           FROM employees e LEFT JOIN departments d ON d.id=e.department_id
                           WHERE e.login_id LIKE %s OR e.first_name LIKE %s OR e.last_name LIKE %s OR e.phone LIKE %s
                           ORDER BY e.first_name LIMIT 50""",(like,like,like,like))
        documents=query("""SELECT ed.id,ed.employee_id,ed.title,ed.document_type,ed.document_number,
                                  CONCAT(e.first_name,' ',COALESCE(e.last_name,'')) employee_name
                           FROM employee_documents ed JOIN employees e ON e.id=ed.employee_id
                           WHERE COALESCE(ed.document_status,'Active')<>'Deleted'
                           AND (ed.title LIKE %s OR ed.document_type LIKE %s OR ed.document_number LIKE %s)
                           ORDER BY ed.uploaded_at DESC LIMIT 50""",(like,like,like))
    return render_template('admin_search.html',q=q,employees=employees,documents=documents)

@app.route('/admin/audit-log')
@admin_required
def admin_audit_log():
    rows=query("""SELECT l.*,a.full_name admin_name,CONCAT(e.first_name,' ',COALESCE(e.last_name,'')) employee_name
                  FROM system_audit_log l LEFT JOIN admins a ON a.id=l.admin_id
                  LEFT JOIN employees e ON e.id=l.employee_id
                  ORDER BY l.created_at DESC LIMIT 500""")
    return render_template('admin_audit_log.html',audit_rows=rows)

@app.route('/admin/system-health')
@admin_required
def admin_system_health():
    allocation_log=os.path.join(os.path.dirname(__file__),'allocation_sync.log')
    backup_root=os.path.join(app.config.get('DATA_ROOT',os.path.dirname(__file__)),'backups')
    health={'db':'Connected','allocation_log_exists':os.path.exists(allocation_log),
            'allocation_log_modified':datetime.fromtimestamp(os.path.getmtime(allocation_log)) if os.path.exists(allocation_log) else None,
            'backup_root':backup_root,'backup_exists':os.path.isdir(backup_root)}
    return render_template('admin_system_health.html',health=health)

@app.route('/admin/employees/<int:eid>/exit',methods=['GET','POST'])
@admin_required
def admin_employee_exit(eid):
    emp=query("SELECT id,login_id,first_name,last_name,status FROM employees WHERE id=%s",(eid,),one=True)
    if not emp:return ('Not found',404)
    if request.method=='POST':
        f=request.form
        query("""INSERT INTO employee_exit_checklist(employee_id,asset_returned,handover_completed,documents_completed,
                 attendance_closed,fnf_status,remarks,updated_by)
                 VALUES(%s,%s,%s,%s,%s,%s,%s,%s)
                 ON DUPLICATE KEY UPDATE asset_returned=VALUES(asset_returned),handover_completed=VALUES(handover_completed),
                 documents_completed=VALUES(documents_completed),attendance_closed=VALUES(attendance_closed),
                 fnf_status=VALUES(fnf_status),remarks=VALUES(remarks),updated_by=VALUES(updated_by)""",
              (eid,1 if f.get('asset_returned') else 0,1 if f.get('handover_completed') else 0,
               1 if f.get('documents_completed') else 0,1 if f.get('attendance_closed') else 0,
               f.get('fnf_status') or 'Pending',(f.get('remarks') or '')[:500],session['admin_id']),commit=True)
        audit_log('UPDATE_EXIT_CHECKLIST','Employee',eid,None,eid)
        flash('Exit/FNF checklist updated.','success')
        return redirect(url_for('admin_employee_exit',eid=eid))
    checklist=query("SELECT * FROM employee_exit_checklist WHERE employee_id=%s",(eid,),one=True) or {}
    return render_template('admin_employee_exit.html',employee=emp,checklist=checklist)

@app.route('/admin/settings',methods=['GET','POST'])
@admin_required
@role_required('Super Admin','Owner')
def admin_settings():
    if request.method=='POST':
        f=request.form
        try:
            query("""UPDATE company_settings SET company_name=%s,company_address=%s,company_phone=%s,company_email=%s,company_website=%s,company_about=%s,owner_name=%s,owner_title=%s,manager_name=%s,manager_title=%s,hr_name=%s,hr_title=%s,management_contact=%s,office_latitude=NULLIF(%s,''),office_longitude=NULLIF(%s,''),office_start=%s,last_reporting=%s,half_day_after=%s,office_end=%s,working_hours=%s,grace_minutes=%s,gps_radius=%s,weekend_days=%s WHERE id=1""",(f.get('company_name'),f.get('company_address'),f.get('company_phone'),f.get('company_email'),f.get('company_website'),f.get('company_about'),f.get('owner_name'),f.get('owner_title'),f.get('manager_name'),f.get('manager_title'),f.get('hr_name'),f.get('hr_title'),f.get('management_contact'),f.get('office_latitude',''),f.get('office_longitude',''),f.get('office_start') or '09:30',f.get('last_reporting') or '09:45',f.get('half_day_after') or '10:30',f.get('office_end') or '18:30',f.get('working_hours') or 9,f.get('grace_minutes') or 15,f.get('gps_radius') or 3,f.get('weekend_days') or 'Sunday'),commit=True)
            logo=request.files.get('company_logo')
            if logo and logo.filename:
                ext=logo.filename.rsplit('.',1)[-1].lower() if '.' in logo.filename else ''
                if ext not in ALLOWED:raise ValueError('Logo must be PNG, JPG, JPEG or WEBP.')
                name=f"company_logo_{int(datetime.now().timestamp())}.{ext}";logo.save(os.path.join(COMPANY_UPLOAD,name));query("UPDATE company_settings SET company_logo=%s WHERE id=1",(name,),commit=True)
            flash('Company settings saved permanently.','success');return redirect(url_for('admin_settings'))
        except Exception as e:flash(f'Unable to save settings: {e}','danger')
    return render_template('admin_settings.html',settings=settings(),edit=request.args.get('edit')=='1',digilocker=dl.configuration_status())


@app.post('/admin/attendance/month-lock')
@admin_required
def admin_attendance_month_lock():
    month=(request.form.get('month') or date.today().strftime('%Y-%m'))[:7]
    action=request.form.get('action') or 'lock'
    note=(request.form.get('note') or '').strip()
    if action=='reopen':
        query("""INSERT INTO attendance_month_locks(month_key,is_locked,reopened_by,reopened_at,note)
                 VALUES(%s,0,%s,NOW(),%s)
                 ON DUPLICATE KEY UPDATE is_locked=0,reopened_by=VALUES(reopened_by),reopened_at=NOW(),note=VALUES(note)""",
              (month,session['admin_id'],note[:500] or None),commit=True)
        audit_log('REOPEN_ATTENDANCE_MONTH','AttendanceMonth',month,note)
        flash(f'Attendance month {month} reopened.','warning')
    else:
        query("""INSERT INTO attendance_month_locks(month_key,is_locked,locked_by,locked_at,note)
                 VALUES(%s,1,%s,NOW(),%s)
                 ON DUPLICATE KEY UPDATE is_locked=1,locked_by=VALUES(locked_by),locked_at=NOW(),note=VALUES(note)""",
              (month,session['admin_id'],note[:500] or None),commit=True)
        audit_log('CLOSE_ATTENDANCE_MONTH','AttendanceMonth',month,note)
        flash(f'Attendance month {month} closed against accidental edits.','success')
    return redirect(url_for('admin_attendance',month=month))

@app.route('/admin/reports')
@admin_required
def admin_reports():
    month=request.args.get('month') or date.today().strftime('%Y-%m')
    report_employee=(request.args.get('employee_id') or '').strip()
    report_department=(request.args.get('department_id') or '').strip()
    try:
        year,mon=map(int,month.split('-'))
        month_start=date(year,mon,1)
        month_end=date(year,mon,calendar.monthrange(year,mon)[1])
    except Exception:
        today=date.today();year=today.year;mon=today.month
        month=today.strftime('%Y-%m');month_start=date(year,mon,1);month_end=date(year,mon,calendar.monthrange(year,mon)[1])

    total_days=calendar.monthrange(year,mon)[1]
    sync_sunday_salary_days(month_start,month_end)
    emp_params=[]
    emp_where="e.status='Active'"
    if report_employee.isdigit():
        emp_where += " AND e.id=%s"; emp_params.append(int(report_employee))
    if report_department.isdigit():
        emp_where += " AND e.department_id=%s"; emp_params.append(int(report_department))
    employees=query(f"""SELECT e.id,e.login_id,e.first_name,e.last_name,e.designation,e.department_id,
                                COALESCE(e.salary,0) salary,d.department_name
                         FROM employees e LEFT JOIN departments d ON d.id=e.department_id
                         WHERE {emp_where} ORDER BY e.first_name,e.last_name""",tuple(emp_params))
    att_rows=query("""SELECT employee_id,attendance_date,status
                      FROM attendance
                      WHERE attendance_date BETWEEN %s AND %s""",(month_start,month_end))
    leave_rows=query("""SELECT employee_id,start_date,end_date,day_type,total_days
                        FROM leave_requests
                        WHERE status='Approved' AND start_date<=%s AND end_date>=%s""",(month_end,month_start))

    by_emp={e['id']:[] for e in employees}
    for row in att_rows: by_emp.setdefault(row['employee_id'],[]).append(row)

    # Approved leave dates are unioned with attendance Leave rows to avoid double-counting full-day leave.
    approved_leave_units={e['id']:0.0 for e in employees}
    for lr in leave_rows:
        s=max(lr['start_date'],month_start); en=min(lr['end_date'],month_end)
        if s>en: continue
        days=(en-s).days+1
        unit=0.5 if str(lr.get('day_type') or '') in ('First Half','Second Half') else 1.0
        approved_leave_units[lr['employee_id']]=approved_leave_units.get(lr['employee_id'],0.0)+(days*unit)

    report_rows=[]
    for e in employees:
        rows=by_emp.get(e['id'],[])
        present=sum(1 for r in rows if r['status']=='Present')
        late=sum(1 for r in rows if r['status']=='Late')
        half_days=sum(1 for r in rows if r['status']=='Half Day')
        holidays=sum(1 for r in rows if r['status']=='Holiday')
        attendance_absent=sum(1 for r in rows if r['status'] in ('Absent','Leave'))
        # Leave requests remain an HR workflow, but payroll/attendance merges them into Absent.
        absent=max(float(attendance_absent),float(approved_leave_units.get(e['id'],0.0)))

        # Confirmed company late rule, reused for any count as combinations of 3 and 2:
        # 3 Late => one full-day deduction across those 3 late days => 2 credited days.
        # 2 Late => one half-day deduction across those 2 late days => 1.5 credited days.
        # 1 remaining Late => 1 credited day.
        groups3,rem=divmod(late,3)
        late_credit=(groups3*2.0)+(1.5 if rem==2 else 1.0 if rem==1 else 0.0)
        final_days=float(present)+float(holidays)+(float(half_days)*0.5)+late_credit
        salary=float(e.get('salary') or 0)
        net_salary=(salary/total_days*final_days) if total_days else 0.0
        report_rows.append({
            'id':e['id'],'employee_name':(f"{e.get('first_name') or ''} {e.get('last_name') or ''}").strip(),
            'login_id':e.get('login_id'),'designation':e.get('designation'),'total_days':total_days,
            'present':present,'late':late,'half_days':half_days,'absent':absent,'final_days':round(final_days,2),
            'salary':salary,'per_day_salary':round((salary/total_days) if total_days else 0,2),'attendance_deduction':round(salary-net_salary,2),'net_salary':round(net_salary,2),'department_name':e.get('department_name')
        })
    departments=query("SELECT id,department_name FROM departments WHERE status='Active' ORDER BY department_name")
    all_employees=query("SELECT id,login_id,first_name,last_name FROM employees WHERE status='Active' ORDER BY first_name,last_name")
    payroll_lock=query("SELECT * FROM payroll_locks WHERE month_key=%s",(month,),one=True)
    return render_template('admin_reports.html',rows=report_rows,month=month,total_days=total_days,departments=departments,
                           all_employees=all_employees,selected_employee_id=report_employee,selected_department_id=report_department,
                           payroll_lock=payroll_lock)


@app.post('/admin/reports/finalize')
@admin_required
@role_required('Super Admin','Owner')
def admin_reports_finalize():
    month=(request.form.get('month') or date.today().strftime('%Y-%m'))[:7]
    if payroll_is_finalized(month):
        flash('Payroll is already finalized for this month.','info')
        return redirect(url_for('admin_reports',month=month))
    year,mon=map(int,month.split('-'))
    start=date(year,mon,1);end=date(year,mon,calendar.monthrange(year,mon)[1])
    employees=query("SELECT id,login_id,first_name,last_name,salary FROM employees WHERE status='Active'")
    att=query("SELECT employee_id,attendance_date,status FROM attendance WHERE attendance_date BETWEEN %s AND %s",(start,end))
    snapshot={'month':month,'employees':employees,'attendance':att}
    query("""INSERT INTO payroll_locks(month_key,is_finalized,finalized_by,finalized_at,snapshot_json)
             VALUES(%s,1,%s,NOW(),%s)
             ON DUPLICATE KEY UPDATE is_finalized=1,finalized_by=VALUES(finalized_by),finalized_at=NOW(),snapshot_json=VALUES(snapshot_json)""",
          (month,session['admin_id'],json.dumps(snapshot,default=str)),commit=True)
    audit_log('FINALIZE_PAYROLL','Payroll',month)
    flash(f'Payroll {month} finalized and locked.','success')
    return redirect(url_for('admin_reports',month=month))

@app.post('/admin/reports/reopen')
@admin_required
@role_required('Super Admin','Owner')
def admin_reports_reopen():
    month=(request.form.get('month') or date.today().strftime('%Y-%m'))[:7]
    query("UPDATE payroll_locks SET is_finalized=0,reopened_by=%s,reopened_at=NOW() WHERE month_key=%s",
          (session['admin_id'],month),commit=True)
    audit_log('REOPEN_PAYROLL','Payroll',month)
    flash(f'Payroll {month} reopened.','warning')
    return redirect(url_for('admin_reports',month=month))

@app.route('/api/employee/dashboard')
@employee_required
def employee_dashboard():
    eid=session['employee_id']
    try: sync_company_attendance_rules()
    except Exception: pass
    try:
        _today=date.today()
        sync_sunday_salary_days(date(_today.year,_today.month,1),date(_today.year,_today.month,calendar.monthrange(_today.year,_today.month)[1]))
    except Exception:
        pass
    emp=query("SELECT e.*,d.department_name FROM employees e LEFT JOIN departments d ON d.id=e.department_id WHERE e.id=%s",(eid,),one=True);att=query("SELECT * FROM attendance WHERE employee_id=%s AND attendance_date=CURDATE()",(eid,),one=True);recent=query("SELECT * FROM attendance WHERE employee_id=%s ORDER BY attendance_date DESC LIMIT 7",(eid,));upcoming=query("SELECT * FROM holidays WHERE status='Active' AND holiday_date>=CURDATE() ORDER BY holiday_date LIMIT 4")
    month=date.today().strftime('%Y-%m');leaderboard=query("""SELECT e.id,e.login_id,CONCAT(e.first_name,' ',COALESCE(e.last_name,'')) employee_name,e.photo,COUNT(a.id) marked,SUM(a.status='Present') on_time,SUM(a.status='Late') late,SUM(a.status='Half Day') half_days,ROUND(COALESCE(SUM(CASE a.status WHEN 'Present' THEN 100 WHEN 'Late' THEN 80 WHEN 'Half Day' THEN 50 WHEN 'Holiday' THEN 100 ELSE 0 END)/NULLIF(COUNT(a.id),0),0),1) score FROM employees e LEFT JOIN attendance a ON a.employee_id=e.id AND DATE_FORMAT(a.attendance_date,'%Y-%m')=%s WHERE e.status='Active' GROUP BY e.id ORDER BY score DESC,on_time DESC,late ASC,employee_name LIMIT 10""",(month,))
    return render_template('employee_dashboard.html',employee=emp,attendance=att,recent=recent,upcoming=upcoming,leaderboard=leaderboard,birthdays=dashboard_birthdays(True))

@app.post('/api/attendance/checkin')
@employee_required
def checkin():
    eid=session['employee_id']
    try: sync_company_attendance_rules()
    except Exception: pass
    existing=query("SELECT * FROM attendance WHERE employee_id=%s AND attendance_date=CURDATE()",(eid,),one=True)
    if existing:
        flash('Attendance is already marked for today.','warning')
        return redirect(url_for('employee_dashboard'))
    s=settings();now=datetime.now()
    if now.date().day!=1 and now.time()>=ATTENDANCE_NO_CHECKIN_CUTOFF:
        query("INSERT IGNORE INTO attendance(employee_id,attendance_date,status,remarks) VALUES(%s,CURDATE(),'Absent','Auto Leave: no check-in by 2:00 PM')",(eid,),commit=True)
        flash('Check-in window closed at 2:00 PM. Today is marked as leave/absent according to company policy.','danger')
        return redirect(url_for('employee_dashboard'))
    report=datetime.combine(now.date(),as_time(s.get('last_reporting'),time(9,45)));late=max(0,int((now-report).total_seconds()//60));half_after=as_time(s.get('half_day_after'),time(10,30));status='Half Day' if now.time()>half_after else ('Late' if late else 'Present');lat=request.form.get('latitude');lng=request.form.get('longitude')
    try:
        configured=s.get('office_latitude') is not None and s.get('office_longitude') is not None
        if configured:
            if not lat or not lng:flash('Location is required for attendance. Allow browser location access and try again.','danger');return redirect(url_for('employee_dashboard'))
            dist=distance_m(lat,lng,s['office_latitude'],s['office_longitude'])
            if dist>int(s.get('gps_radius') or 3):flash(f'Check-in blocked: you are approximately {dist:.0f} m away from office.','danger');return redirect(url_for('employee_dashboard'))
        query("INSERT INTO attendance(employee_id,attendance_date,check_in,check_in_lat,check_in_lng,late_minutes,status) VALUES(%s,CURDATE(),NOW(),NULLIF(%s,''),NULLIF(%s,''),%s,%s)",(eid,lat or '',lng or '',late,status),commit=True);flash('Check-in successful.'+(' Half day has been marked because check-in was after 10:30 AM.' if status=='Half Day' else ''),'success')
    except Exception as e:flash(f'Check-in could not be completed: {e}','danger')
    return redirect(url_for('employee_dashboard'))

@app.post('/api/attendance/late-reason')
@employee_required
def late_reason():
    eid=session['employee_id'];a=query("SELECT * FROM attendance WHERE employee_id=%s AND attendance_date=CURDATE()",(eid,),one=True);reason=(request.form.get('late_reason') or '').strip()
    if not a or int(a.get('late_minutes') or 0)<=0:return redirect(url_for('employee_dashboard'))
    if len(reason)<3:flash('Please enter the reason for being late.','danger');return redirect(url_for('employee_dashboard'))
    query("UPDATE attendance SET late_reason=%s WHERE id=%s",(reason[:500],a['id']),commit=True);flash('Late-arrival reason saved.','success');return redirect(url_for('employee_dashboard'))

@app.post('/api/attendance/checkout')
@employee_required
def checkout():
    eid=session['employee_id'];a=query("SELECT * FROM attendance WHERE employee_id=%s AND attendance_date=CURDATE()",(eid,),one=True)
    if not a or not a['check_in'] or a['check_out']:flash('No active check-in found.','warning');return redirect(url_for('employee_dashboard'))
    if int(a.get('late_minutes') or 0)>0 and not (a.get('late_reason') or '').strip():flash('Submit your late-arrival reason before checking out.','warning');return redirect(url_for('employee_dashboard'))
    s=settings();now=datetime.now();working=int((now-a['check_in']).total_seconds()//60);end=datetime.combine(now.date(),as_time(s.get('office_end'),time(18,30)));overtime=max(0,int((now-end).total_seconds()//60));early=max(0,int((end-now).total_seconds()//60));half_after=as_time(s.get('half_day_after'),time(10,30));final_status='Half Day' if (a['check_in'].time()>half_after or now.time()<ATTENDANCE_EARLY_CHECKOUT_CUTOFF or a.get('status')=='Half Day') else a['status']
    query("UPDATE attendance SET check_out=NOW(),check_out_lat=NULLIF(%s,''),check_out_lng=NULLIF(%s,''),working_minutes=%s,overtime_minutes=%s,early_exit_minutes=%s,status=%s WHERE id=%s",(request.form.get('latitude',''),request.form.get('longitude',''),working,overtime,early,final_status,a['id']),commit=True);flash('Check-out successful.'+(' Half day remains marked.' if final_status=='Half Day' else ''),'success');return redirect(url_for('employee_dashboard'))


@app.route('/employee/feedback',methods=['GET','POST'])
@employee_required
def employee_feedback():
    eid=session['employee_id']
    if request.method=='POST':
        message=(request.form.get('message') or '').strip()
        if len(message)<5:
            flash('Please enter a little more detail in your feedback.','warning')
        else:
            query("INSERT INTO employee_feedback(employee_id,message) VALUES(%s,%s)",(eid,message[:4000]),commit=True)
            flash('Feedback submitted successfully.','success')
        return redirect(url_for('employee_feedback'))
    rows=query("SELECT * FROM employee_feedback WHERE employee_id=%s ORDER BY created_at DESC",(eid,))
    return render_template('employee_feedback.html',feedback_rows=rows)

@app.route('/admin/feedback')
@admin_required
def admin_feedback():
    status=(request.args.get('status') or '').strip()
    params=[];where='1=1'
    if status in ('New','Reviewed','Resolved'):
        where='f.status=%s';params=[status]
    rows=query(f"""SELECT f.*,e.login_id,CONCAT(e.first_name,' ',COALESCE(e.last_name,'')) employee_name
                    FROM employee_feedback f JOIN employees e ON e.id=f.employee_id
                    WHERE {where} ORDER BY (f.status='New') DESC,f.created_at DESC""",tuple(params))
    return render_template('admin_feedback.html',feedback_rows=rows,status_filter=status)

@app.post('/admin/feedback/<int:fid>/update')
@admin_required
def admin_feedback_update(fid):
    status=request.form.get('status') or 'Reviewed'
    if status not in ('New','Reviewed','Resolved'): status='Reviewed'
    reply=(request.form.get('admin_reply') or '').strip()
    fb=query("SELECT * FROM employee_feedback WHERE id=%s",(fid,),one=True)
    if not fb:
        flash('Feedback record not found.','danger'); return redirect(url_for('admin_feedback'))
    if int(fb.get('response_locked') or 0)==1:
        flash('This feedback response is locked after saving and is now view-only.','warning')
        return redirect(url_for('admin_feedback'))
    query("UPDATE employee_feedback SET status=%s,admin_reply=%s,reviewed_by=%s,reviewed_at=NOW(),response_locked=1,locked_at=NOW() WHERE id=%s",
          (status,reply[:4000] or None,session['admin_id'],fid),commit=True)
    notify_employee(fb['employee_id'],'Feedback updated',f'Your feedback status is now {status}.'+(' Admin has also replied.' if reply else ''),'Feedback')
    audit_log('UPDATE_FEEDBACK_LOCKED','Feedback',fid,{'status':status,'response_locked':True},fb['employee_id'])
    flash('Feedback response saved and locked. It is now view-only.','success')
    return redirect(url_for('admin_feedback'))

def digital_id_serializer():
    return URLSafeSerializer(app.config['SECRET_KEY'],salt='grsj-digital-id-v1')

def employee_verification_token(emp):
    return digital_id_serializer().dumps({'employee_id':int(emp['id']),'login_id':emp['login_id']})

def digital_id_qr_drawing(value,size=32*mm):
    qr=QrCodeWidget(value)
    x1,y1,x2,y2=qr.getBounds(); qw=x2-x1; qh=y2-y1
    d=Drawing(size,size,transform=[size/qw,0,0,size/qh,0,0])
    d.add(qr)
    return d

def digital_id_image_path(employee):
    name=(employee.get('photo') or '').strip() if employee else ''
    if name:
        path=os.path.join(app.config['UPLOAD_FOLDER'],name)
        if os.path.isfile(path): return path
    fallback=os.path.join(app.static_folder,'images','default_user.png')
    return fallback if os.path.isfile(fallback) else None

def digital_id_logo_path(company):
    name=(company.get('company_logo') or '').strip() if company else ''
    if name:
        path=os.path.join(COMPANY_UPLOAD,name)
        if os.path.isfile(path): return path
    fallback=os.path.join(app.static_folder,'images','Logo.png')
    return fallback if os.path.isfile(fallback) else None

def digital_id_valid_till(joining_date):
    if not joining_date:
        return None
    try:
        return joining_date.replace(year=joining_date.year + 1)
    except ValueError:
        return joining_date.replace(year=joining_date.year + 1, day=28)

@app.route('/employee/digital-id')
@employee_required
def employee_digital_id():
    eid=session['employee_id']
    emp=query("SELECT e.*,d.department_name FROM employees e LEFT JOIN departments d ON d.id=e.department_id WHERE e.id=%s",(eid,),one=True)
    token=employee_verification_token(emp)
    verify_url=url_for('verify_employee_id',token=token,_external=True)
    return render_template('employee_digital_id.html',employee=emp,verify_url=verify_url,
                           valid_till=digital_id_valid_till(emp.get('joining_date')))

@app.route('/employee/digital-id/qr.svg')
@employee_required
def employee_digital_id_qr():
    emp=query("SELECT id,login_id FROM employees WHERE id=%s",(session['employee_id'],),one=True)
    if not emp: return Response(status=404)
    verify_url=url_for('verify_employee_id',token=employee_verification_token(emp),_external=True)
    svg=renderSVG.drawToString(digital_id_qr_drawing(verify_url,34*mm))
    return Response(svg,mimetype='image/svg+xml',headers={'Cache-Control':'private, no-store'})

@app.route('/verify/employee/<token>')
def verify_employee_id(token):
    try:
        payload=digital_id_serializer().loads(token)
        eid=int(payload.get('employee_id'))
        login_id=str(payload.get('login_id') or '')
    except (BadSignature,ValueError,TypeError):
        return render_template('employee_id_verify.html',verified=False,employee=None),404
    emp=query("""SELECT e.id,e.employee_id,e.login_id,e.first_name,e.last_name,e.designation,e.status,e.employment_stage,
                        d.department_name
                 FROM employees e LEFT JOIN departments d ON d.id=e.department_id
                 WHERE e.id=%s AND e.login_id=%s""",(eid,login_id),one=True)
    valid=bool(emp and emp.get('status')=='Active' and (emp.get('employment_stage') or 'Active') not in ('Terminated','Archived','Resigned'))
    return render_template('employee_id_verify.html',verified=valid,employee=emp),200 if valid else 404

@app.route('/verify/employee-id/<employee_code>')
def verify_employee_by_public_id(employee_code):
    code=(employee_code or '').strip().upper()
    if not code or len(code)>40 or not re.fullmatch(r'[A-Z0-9_-]+',code):
        return render_template('employee_id_verify.html',verified=False,employee=None),404
    emp=query("""SELECT e.id,e.employee_id,e.login_id,e.first_name,e.last_name,e.designation,e.status,e.employment_stage,
                        d.department_name
                 FROM employees e LEFT JOIN departments d ON d.id=e.department_id
                 WHERE UPPER(COALESCE(e.employee_id,''))=%s OR UPPER(COALESCE(e.login_id,''))=%s
                 LIMIT 1""",(code,code),one=True)
    valid=bool(emp and emp.get('status')=='Active' and (emp.get('employment_stage') or 'Active') not in ('Terminated','Archived','Resigned'))
    return render_template('employee_id_verify.html',verified=valid,employee=emp),200 if valid else 404

@app.route('/employee/digital-id/download')
@employee_required
def employee_digital_id_download():
    emp=query("SELECT e.*,d.department_name FROM employees e LEFT JOIN departments d ON d.id=e.department_id WHERE e.id=%s",(session['employee_id'],),one=True)
    if not emp:
        flash('Employee record not found.','danger')
        return redirect(url_for('employee_profile'))
    company=settings()
    token=employee_verification_token(emp)
    verify_url=url_for('verify_employee_id',token=token,_external=True)
    logo_path=digital_id_logo_path(company)
    photo_path=digital_id_image_path(emp)
    qr_svg=renderSVG.drawToString(digital_id_qr_drawing(verify_url,34*mm))
    # app.css is now a stable modular entrypoint. For the self-contained PDF
    # template, concatenate the ordered CSS parts directly because @import URLs
    # are not resolvable inside Playwright page.set_content().
    css_parts_dir=os.path.join(app.static_folder,'css','parts')
    css_files=[]
    if os.path.isdir(css_parts_dir):
        css_files=sorted(
            os.path.join(css_parts_dir,name)
            for name in os.listdir(css_parts_dir)
            if name.lower().endswith('.css')
        )
    if css_files:
        css_text='\n'.join(open(path,'r',encoding='utf-8').read() for path in css_files)
    else:
        css_path=os.path.join(app.static_folder,'css','app.css')
        with open(css_path,'r',encoding='utf-8') as fh:
            css_text=fh.read()
    rendered=render_template(
        'employee_digital_id_print.html', employee=emp, company=company,
        valid_till=digital_id_valid_till(emp.get('joining_date')),
        css_text=css_text,
        card_logo_src=file_to_data_uri(logo_path) if logo_path else '',
        card_photo_src=file_to_data_uri(photo_path) if photo_path else '',
        card_qr_src=svg_to_data_uri(qr_svg),
        verify_url=verify_url
    )
    try:
        out=build_digital_id_pdf(rendered)
    except Exception as e:
        app.logger.exception('Digital ID browser PDF generation failed')
        flash(f'Unable to generate Digital ID PDF: {e}','danger')
        return redirect(url_for('employee_digital_id'))
    filename=f"GRSJ_Digital_ID_{re.sub(r'[^A-Za-z0-9_-]+','_',str(emp.get('employee_id') or emp.get('login_id') or emp.get('id')))}.pdf"
    return send_file(out,mimetype='application/pdf',as_attachment=True,download_name=filename)

def offer_letter_form_data(emp, company, form):
    document_date=(form.get('document_date') or date.today().isoformat()).strip()
    try: issue_dt=datetime.strptime(document_date,'%Y-%m-%d').date()
    except ValueError: issue_dt=date.today(); document_date=issue_dt.isoformat()
    acceptance_date=(form.get('acceptance_date') or (issue_dt+timedelta(days=7)).isoformat()).strip()
    joining_raw=(form.get('joining_date') or (emp.get('joining_date').isoformat() if emp.get('joining_date') else '')).strip()
    monthly=(form.get('monthly_salary') or emp.get('salary') or 0)
    ref=(form.get('reference_no') or '').strip()
    if not ref:
        ref=f"GRSJ/OL/{issue_dt.strftime('%Y%m')}/{emp.get('login_id') or emp.get('id')}"
    company_address=company.get('company_address') or 'Property No 8, Kh No 108, Mayur Vihar Phase 1, Harijan Basti Jagat Colony, Chilla Saroda Bangar, New Delhi - 110091'
    return {
        'employee_id':emp['id'],
        'employee_name':f"{emp.get('first_name') or ''} {emp.get('last_name') or ''}".strip(),
        'login_id':emp.get('login_id') or '',
        'designation':(form.get('designation') or emp.get('designation') or 'Employee').strip(),
        'department':(form.get('department') or emp.get('department_name') or emp.get('department') or 'General').strip(),
        'joining_date':joining_raw,
        'monthly_salary':monthly,
        'document_date':document_date,
        'acceptance_date':acceptance_date,
        'reference_no':ref,
        'probation_months':(form.get('probation_months') or '3').strip(),
        'work_location':(form.get('work_location') or company_address).strip(),
        'reporting_to':(form.get('reporting_to') or '').strip(),
        'additional_terms':(form.get('additional_terms') or form.get('notes') or '').strip(),
        'company_name':company.get('company_name') or 'Guru Ram Singh Ji Associates',
    }


def appointment_letter_form_data(emp, company, form):
    document_date=(form.get('document_date') or date.today().isoformat()).strip()
    try: issue_dt=datetime.strptime(document_date,'%Y-%m-%d').date()
    except ValueError: issue_dt=date.today(); document_date=issue_dt.isoformat()
    joining_raw=(form.get('joining_date') or (emp.get('joining_date').isoformat() if emp.get('joining_date') else '')).strip()
    effective=(form.get('effective_date') or joining_raw or document_date).strip()
    ref=(form.get('reference_no') or '').strip()
    if not ref: ref=f"GRSJ/AL/{issue_dt.strftime('%Y%m')}/{emp.get('login_id') or emp.get('id')}"
    company_address=company.get('company_address') or 'Property No 8, Kh No 108, Mayur Vihar Phase 1, Harijan Basti Jagat Colony, Chilla Saroda Bangar, New Delhi - 110091'
    return {'employee_id':emp['id'],'employee_name':f"{emp.get('first_name') or ''} {emp.get('last_name') or ''}".strip(),
      'login_id':emp.get('login_id') or '','designation':(form.get('designation') or emp.get('designation') or 'Employee').strip(),
      'department':(form.get('department') or emp.get('department_name') or emp.get('department') or 'General').strip(),
      'joining_date':joining_raw,'effective_date':effective,'document_date':document_date,'reference_no':ref,
      'probation_months':(form.get('probation_months') or '3').strip(),'work_location':(form.get('work_location') or company_address).strip(),
      'reporting_to':(form.get('reporting_to') or '').strip(),'additional_terms':(form.get('additional_terms') or '').strip(),
      'company_name':company.get('company_name') or 'Guru Ram Singh Ji Associates'}



def termination_letter_form_data(emp, company, form):
    document_date=(form.get('document_date') or date.today().isoformat()).strip()
    try: issue_dt=datetime.strptime(document_date,'%Y-%m-%d').date()
    except ValueError: issue_dt=date.today(); document_date=issue_dt.isoformat()
    effective=(form.get('effective_date') or document_date).strip()
    last_working=(form.get('last_working_date') or effective).strip()
    ref=(form.get('reference_no') or '').strip()
    if not ref: ref=f"GRSJ/TL/{issue_dt.strftime('%Y%m')}/{emp.get('login_id') or emp.get('id')}"
    return {
      'employee_id':emp['id'],'employee_name':f"{emp.get('first_name') or ''} {emp.get('last_name') or ''}".strip(),
      'login_id':emp.get('login_id') or '','designation':emp.get('designation') or 'Employee',
      'department':emp.get('department_name') or emp.get('department') or 'General',
      'document_date':document_date,'effective_date':effective,'last_working_date':last_working,'reference_no':ref,
      'termination_category':(form.get('termination_category') or 'Company Initiated').strip(),
      'termination_reason':(form.get('termination_reason') or '').strip(),
      'notice_details':(form.get('notice_details') or 'As per applicable company policy').strip(),
      'asset_instructions':(form.get('asset_instructions') or 'Return all company property, records, credentials and assets in your possession on or before the last working day.').strip(),
      'fnf_details':(form.get('fnf_details') or "In accordance with the company's applicable policies and the law, the 'Full and Final settlement' process will be completed only after the necessary clearances have been obtained and a period of 45 days has elapsed.").strip(),
      'additional_terms':(form.get('additional_terms') or '').strip(),
      'company_name':company.get('company_name') or 'Guru Ram Singh Ji Associates'}



def promotion_letter_form_data(emp, company, form):
    document_date=(form.get('document_date') or date.today().isoformat()).strip()
    try: issue_dt=datetime.strptime(document_date,'%Y-%m-%d').date()
    except ValueError: issue_dt=date.today(); document_date=issue_dt.isoformat()
    effective=(form.get('effective_date') or document_date).strip()
    ref=(form.get('reference_no') or '').strip()
    if not ref: ref=f"GRSJ/PL/{issue_dt.strftime('%Y%m')}/{emp.get('login_id') or emp.get('id')}"
    return {
      'employee_id':emp['id'],'employee_name':f"{emp.get('first_name') or ''} {emp.get('last_name') or ''}".strip(),
      'login_id':emp.get('login_id') or '',
      'current_designation':(form.get('current_designation') or emp.get('designation') or 'Employee').strip(),
      'new_designation':(form.get('new_designation') or '').strip(),
      'current_department':(form.get('current_department') or emp.get('department_name') or emp.get('department') or 'General').strip(),
      'new_department':(form.get('new_department') or emp.get('department_name') or emp.get('department') or 'General').strip(),
      'document_date':document_date,'effective_date':effective,'reference_no':ref,
      'reporting_to':(form.get('reporting_to') or '').strip(),
      'salary_revision_applicable':(form.get('salary_revision_applicable') or 'No').strip(),
      'revised_salary':(form.get('revised_salary') or '').strip(),
      'revised_responsibilities':(form.get('revised_responsibilities') or '').strip(),
      'additional_terms':(form.get('additional_terms') or '').strip(),
      'company_name':company.get('company_name') or 'Guru Ram Singh Ji Associates'}



def pay_slip_form_data(emp, company, form):
    salary_month=(form.get('salary_month') or date.today().strftime('%Y-%m')).strip()
    try:
        year,mon=map(int,salary_month.split('-'))
        month_start=date(year,mon,1)
        month_end=date(year,mon,calendar.monthrange(year,mon)[1])
    except Exception:
        today=date.today(); year=today.year; mon=today.month
        salary_month=today.strftime('%Y-%m')
        month_start=date(year,mon,1); month_end=date(year,mon,calendar.monthrange(year,mon)[1])

    total_days=calendar.monthrange(year,mon)[1]
    sync_sunday_salary_days(month_start,month_end)

    rows=query("""SELECT attendance_date,status FROM attendance
                  WHERE employee_id=%s AND attendance_date BETWEEN %s AND %s""",
               (emp['id'],month_start,month_end))
    leave_rows=query("""SELECT start_date,end_date,day_type,total_days FROM leave_requests
                        WHERE employee_id=%s AND status='Approved' AND start_date<=%s AND end_date>=%s""",
                     (emp['id'],month_end,month_start))

    present=sum(1 for r in rows if r['status']=='Present')
    late=sum(1 for r in rows if r['status']=='Late')
    half_days=sum(1 for r in rows if r['status']=='Half Day')
    holidays=sum(1 for r in rows if r['status']=='Holiday')
    attendance_absent=sum(1 for r in rows if r['status'] in ('Absent','Leave'))

    approved_leave_units=0.0
    for lr in leave_rows:
        s=max(lr['start_date'],month_start); en=min(lr['end_date'],month_end)
        if s>en: continue
        days=(en-s).days+1
        unit=0.5 if str(lr.get('day_type') or '') in ('First Half','Second Half') else 1.0
        approved_leave_units += days*unit
    absent=max(float(attendance_absent),float(approved_leave_units))

    groups3,rem=divmod(late,3)
    late_credit=(groups3*2.0)+(1.5 if rem==2 else 1.0 if rem==1 else 0.0)
    final_days=float(present)+float(holidays)+(float(half_days)*0.5)+late_credit

    monthly_salary=float(emp.get('salary') or 0)
    earned_salary=(monthly_salary/total_days*final_days) if total_days else 0.0
    incentive=float(form.get('incentive') or 0)
    other_deductions=float(form.get('other_deductions') or 0)

    document_date=(form.get('document_date') or date.today().isoformat()).strip()
    ref=(form.get('reference_no') or '').strip()
    if not ref:
        ref=f"GRSJ/PS/{year}{mon:02d}/{emp.get('login_id') or emp.get('id')}"

    return {
      'employee_id':emp['id'],
      'employee_name':f"{emp.get('first_name') or ''} {emp.get('last_name') or ''}".strip(),
      'login_id':emp.get('login_id') or '',
      'designation':emp.get('designation') or 'Employee',
      'department':emp.get('department_name') or emp.get('department') or 'General',
      'salary_month':salary_month,'document_date':document_date,'reference_no':ref,
      'total_days':total_days,'present':present,'late':late,'half_days':half_days,
      'holidays':holidays,'absent':round(absent,2),'final_days':round(final_days,2),
      'monthly_salary':round(monthly_salary,2),'earned_salary':round(earned_salary,2),
      'incentive':round(incentive,2),'other_deductions':round(other_deductions,2),
      'remarks':(form.get('remarks') or '').strip(),
      'company_name':company.get('company_name') or 'Guru Ram Singh Ji Associates'
    }



def experience_letter_form_data(emp, company, form):
    document_date=(form.get('document_date') or date.today().isoformat()).strip()
    try:
        issue_dt=datetime.strptime(document_date,'%Y-%m-%d').date()
    except ValueError:
        issue_dt=date.today(); document_date=issue_dt.isoformat()

    joining_raw=(form.get('joining_date') or (emp.get('joining_date').isoformat() if emp.get('joining_date') else '')).strip()
    last_working=(form.get('last_working_date') or date.today().isoformat()).strip()
    ref=(form.get('reference_no') or '').strip()
    if not ref:
        ref=f"GRSJ/EL/{issue_dt.strftime('%Y%m')}/{emp.get('login_id') or emp.get('id')}"

    return {
      'employee_id':emp['id'],
      'employee_name':f"{emp.get('first_name') or ''} {emp.get('last_name') or ''}".strip(),
      'login_id':emp.get('login_id') or '',
      'designation':(form.get('designation') or emp.get('designation') or 'Employee').strip(),
      'department':(form.get('department') or emp.get('department_name') or emp.get('department') or 'General').strip(),
      'joining_date':joining_raw,
      'last_working_date':last_working,
      'document_date':document_date,
      'reference_no':ref,
      'role_summary':(form.get('role_summary') or '').strip(),
      'conduct_remarks':(form.get('conduct_remarks') or '').strip(),
      'additional_remarks':(form.get('additional_remarks') or '').strip(),
      'company_name':company.get('company_name') or 'Guru Ram Singh Ji Associates'
    }



def pip_form_data(emp,company,form):
    document_date=(form.get('document_date') or date.today().isoformat()).strip()
    try: issue_dt=datetime.strptime(document_date,'%Y-%m-%d').date()
    except ValueError: issue_dt=date.today(); document_date=issue_dt.isoformat()
    ref=(form.get('reference_no') or '').strip() or f"GRSJ/PIP/{issue_dt.strftime('%Y%m')}/{emp.get('login_id') or emp.get('id')}"
    return {
      'employee_id':emp['id'],'employee_name':f"{emp.get('first_name') or ''} {emp.get('last_name') or ''}".strip(),
      'login_id':emp.get('login_id') or '','designation':emp.get('designation') or 'Employee',
      'department':emp.get('department_name') or emp.get('department') or 'General','document_date':document_date,'reference_no':ref,
      'period_start':(form.get('period_start') or '').strip(),'period_end':(form.get('period_end') or '').strip(),
      'review_frequency':(form.get('review_frequency') or 'Weekly').strip(),'review_date':(form.get('review_date') or '').strip(),
      'performance_concerns':(form.get('performance_concerns') or '').strip(),'expectations':(form.get('expectations') or '').strip(),
      'action_plan':(form.get('action_plan') or '').strip(),'management_support':(form.get('management_support') or '').strip(),
      'consequence':(form.get('consequence') or '').strip(),'additional_remarks':(form.get('additional_remarks') or '').strip(),
      'company_name':company.get('company_name') or 'Guru Ram Singh Ji Associates'
    }

def warning_letter_form_data(emp,company,form):
    document_date=(form.get('document_date') or date.today().isoformat()).strip()
    try: issue_dt=datetime.strptime(document_date,'%Y-%m-%d').date()
    except ValueError: issue_dt=date.today();document_date=issue_dt.isoformat()
    ref=(form.get('reference_no') or '').strip() or f"GRSJ/WL/{issue_dt.strftime('%Y%m')}/{emp.get('login_id') or emp.get('id')}"
    return {'employee_id':emp['id'],'employee_name':f"{emp.get('first_name') or ''} {emp.get('last_name') or ''}".strip(),
      'login_id':emp.get('login_id') or '','designation':emp.get('designation') or 'Employee',
      'department':emp.get('department_name') or emp.get('department') or 'General','document_date':document_date,'reference_no':ref,
      'warning_level':(form.get('warning_level') or 'Written Warning').strip(),'warning_subject':(form.get('warning_subject') or '').strip(),
      'incident_date':(form.get('incident_date') or document_date).strip(),'incident_details':(form.get('incident_details') or '').strip(),
      'corrective_action':(form.get('corrective_action') or '').strip(),'improvement_deadline':(form.get('improvement_deadline') or '').strip(),
      'consequence':(form.get('consequence') or '').strip(),'additional_remarks':(form.get('additional_remarks') or '').strip(),
      'company_name':company.get('company_name') or 'Guru Ram Singh Ji Associates'}



def relieving_letter_form_data(emp,company,form):
    document_date=(form.get('document_date') or date.today().isoformat()).strip()
    try: issue_dt=datetime.strptime(document_date,'%Y-%m-%d').date()
    except ValueError: issue_dt=date.today();document_date=issue_dt.isoformat()
    joining=(form.get('joining_date') or (emp.get('joining_date').isoformat() if emp.get('joining_date') else '')).strip()
    last=(form.get('last_working_date') or document_date).strip()
    ref=(form.get('reference_no') or '').strip() or f"GRSJ/RL/{issue_dt.strftime('%Y%m')}/{emp.get('login_id') or emp.get('id')}"
    return {'employee_id':emp['id'],'employee_name':f"{emp.get('first_name') or ''} {emp.get('last_name') or ''}".strip(),
      'login_id':emp.get('login_id') or '','designation':(form.get('designation') or emp.get('designation') or 'Employee').strip(),
      'department':(form.get('department') or emp.get('department_name') or emp.get('department') or 'General').strip(),
      'joining_date':joining,'last_working_date':last,'document_date':document_date,'reference_no':ref,
      'separation_type':(form.get('separation_type') or 'Resignation').strip(),'clearance_status':(form.get('clearance_status') or 'Completed').strip(),
      'handover_remarks':(form.get('handover_remarks') or '').strip(),'fnf_status':(form.get('fnf_status') or '').strip(),
      'additional_remarks':(form.get('additional_remarks') or '').strip(),'company_name':company.get('company_name') or 'Guru Ram Singh Ji Associates'}


@app.route('/admin/documents' ,methods=['GET','POST'])
@admin_required
def admin_hr_documents():
    employees=query("SELECT id,login_id,first_name,last_name,designation,joining_date,salary,status FROM employees ORDER BY (status='Active') DESC,first_name,last_name")
    selected_employee=None
    eid=(request.values.get('employee_id') or '').strip()
    if eid.isdigit():
        selected_employee=query("""SELECT e.*,d.department_name FROM employees e LEFT JOIN departments d ON d.id=e.department_id WHERE e.id=%s""",(int(eid),),one=True)
    selected_type=request.values.get('document_type') or ''
    company=settings()
    default_acceptance=(date.today()+timedelta(days=7)).isoformat()
    if request.method=='POST':
        dtype=(request.form.get('document_type') or '').strip()
        if not dtype or not selected_employee:
            flash('Select a document type and employee.','warning')
        elif dtype=='Offer Letter':
            data=offer_letter_form_data(selected_employee,company,request.form)
            background=os.path.join(app.static_folder,'images','grsj_offer_letterhead_clean.png')
            pdf=build_offer_letter_pdf(data,background)
            action=(request.form.get('action') or 'preview').strip().lower()
            safe_id=re.sub(r'[^A-Za-z0-9_-]+','_',str(selected_employee.get('login_id') or selected_employee['id']))
            filename=f"GRSJ_Offer_Letter_{safe_id}_{data['document_date'].replace('-','')}.pdf"
            if action=='preview':
                return send_file(pdf,mimetype='application/pdf',as_attachment=False,download_name=filename)
            disk_name=f"hr_{selected_employee['id']}_{int(datetime.now().timestamp())}_{filename}"[:255]
            disk_path=os.path.join(DOCUMENT_UPLOAD,disk_name)
            with open(disk_path,'wb') as f:f.write(pdf.getvalue())
            fields=json.dumps(data,ensure_ascii=False,default=str)
            query("""INSERT INTO hr_document_drafts(employee_id,document_type,document_date,reference_no,notes,fields_json,status,generated_file,created_by)
                     VALUES(%s,'Offer Letter',%s,%s,%s,%s,'Generated',%s,%s)""",
                  (selected_employee['id'],data['document_date'],data['reference_no'],data['additional_terms'] or None,fields,disk_name,session['admin_id']),commit=True)
            query("""INSERT INTO employee_documents(employee_id,title,document_type,file_name,original_name,verification_status,verification_source,verified_at)
                     VALUES(%s,%s,'Offer Letter',%s,%s,'Verified','HRMS Generated',NOW())""",
                  (selected_employee['id'],f"Offer Letter - {data['document_date']}",disk_name,filename),commit=True)
            notify_employee(selected_employee['id'],'New Offer Letter','Offer Letter has been issued and is available in your Documents section.','Document')
            audit_log('ISSUE_HR_DOCUMENT','EmployeeDocument',None,{'type':'Offer Letter'},selected_employee['id'])
            flash('Offer Letter generated and issued successfully. It is now available in the employee Documents section.','success')
            return redirect(url_for('admin_hr_documents',employee_id=selected_employee['id'],document_type='Offer Letter'))
        elif dtype=='Appointment Letter':
            data=appointment_letter_form_data(selected_employee,company,request.form)
            background=os.path.join(app.static_folder,'images','grsj_offer_letterhead_clean.png')
            pdf=build_appointment_letter_pdf(data,background)
            action=(request.form.get('action') or 'preview').strip().lower()
            safe_id=re.sub(r'[^A-Za-z0-9_-]+','_',str(selected_employee.get('login_id') or selected_employee['id']))
            filename=f"GRSJ_Appointment_Letter_{safe_id}_{data['document_date'].replace('-','')}.pdf"
            if action=='preview':
                return send_file(pdf,mimetype='application/pdf',as_attachment=False,download_name=filename)
            disk_name=f"hr_{selected_employee['id']}_{int(datetime.now().timestamp())}_{filename}"[:255]
            disk_path=os.path.join(DOCUMENT_UPLOAD,disk_name)
            with open(disk_path,'wb') as f:f.write(pdf.getvalue())
            fields=json.dumps(data,ensure_ascii=False,default=str)
            query("INSERT INTO hr_document_drafts(employee_id,document_type,document_date,reference_no,notes,fields_json,status,generated_file,created_by) VALUES(%s,'Appointment Letter',%s,%s,%s,%s,'Generated',%s,%s)",
                  (selected_employee['id'],data['document_date'],data['reference_no'],data['additional_terms'] or None,fields,disk_name,session['admin_id']),commit=True)
            query("INSERT INTO employee_documents(employee_id,title,document_type,file_name,original_name,verification_status,verification_source,verified_at) VALUES(%s,%s,'Appointment Letter',%s,%s,'Verified','HRMS Generated',NOW())",
                  (selected_employee['id'],f"Appointment Letter - {data['document_date']}",disk_name,filename),commit=True)
            notify_employee(selected_employee['id'],'New Appointment Letter','Appointment Letter has been issued and is available in your Documents section.','Document')
            audit_log('ISSUE_HR_DOCUMENT','EmployeeDocument',None,{'type':'Appointment Letter'},selected_employee['id'])
            flash('Appointment Letter generated and issued successfully. It is now available in the employee Documents section.','success')
            return redirect(url_for('admin_hr_documents',employee_id=selected_employee['id'],document_type='Appointment Letter'))
        elif dtype=='Termination Letter':
            data=termination_letter_form_data(selected_employee,company,request.form)
            background=os.path.join(app.static_folder,'images','grsj_offer_letterhead_clean.png')
            pdf=build_termination_letter_pdf(data,background)
            action=(request.form.get('action') or 'preview').strip().lower()
            safe_id=re.sub(r'[^A-Za-z0-9_-]+','_',str(selected_employee.get('login_id') or selected_employee['id']))
            filename=f"GRSJ_Termination_Letter_{safe_id}_{data['document_date'].replace('-','')}.pdf"
            if action=='preview':
                return send_file(pdf,mimetype='application/pdf',as_attachment=False,download_name=filename)
            disk_name=f"hr_{selected_employee['id']}_{int(datetime.now().timestamp())}_{filename}"[:255]
            disk_path=os.path.join(DOCUMENT_UPLOAD,disk_name)
            with open(disk_path,'wb') as f:f.write(pdf.getvalue())
            fields=json.dumps(data,ensure_ascii=False,default=str)
            query("INSERT INTO hr_document_drafts(employee_id,document_type,document_date,reference_no,notes,fields_json,status,generated_file,created_by) VALUES(%s,'Termination Letter',%s,%s,%s,%s,'Generated',%s,%s)",
                  (selected_employee['id'],data['document_date'],data['reference_no'],data['termination_reason'],fields,disk_name,session['admin_id']),commit=True)
            query("INSERT INTO employee_documents(employee_id,title,document_type,file_name,original_name,verification_status,verification_source,verified_at) VALUES(%s,%s,'Termination Letter',%s,%s,'Verified','HRMS Generated',NOW())",
                  (selected_employee['id'],f"Termination Letter - {data['document_date']}",disk_name,filename),commit=True)
            notify_employee(selected_employee['id'],'New Termination Letter','Termination Letter has been issued and is available in your Documents section.','Document')
            audit_log('ISSUE_HR_DOCUMENT','EmployeeDocument',None,{'type':'Termination Letter'},selected_employee['id'])
            flash('Termination Letter generated and issued successfully. It is now available in the employee Documents section.','success')
            return redirect(url_for('admin_hr_documents',employee_id=selected_employee['id'],document_type='Termination Letter'))
        elif dtype=='Promotion Letter':
            data=promotion_letter_form_data(selected_employee,company,request.form)
            background=os.path.join(app.static_folder,'images','grsj_offer_letterhead_clean.png')
            pdf=build_promotion_letter_pdf(data,background)
            action=(request.form.get('action') or 'preview').strip().lower()
            safe_id=re.sub(r'[^A-Za-z0-9_-]+','_',str(selected_employee.get('login_id') or selected_employee['id']))
            filename=f"GRSJ_Promotion_Letter_{safe_id}_{data['document_date'].replace('-','')}.pdf"
            if action=='preview':
                return send_file(pdf,mimetype='application/pdf',as_attachment=False,download_name=filename)
            disk_name=f"hr_{selected_employee['id']}_{int(datetime.now().timestamp())}_{filename}"[:255]
            disk_path=os.path.join(DOCUMENT_UPLOAD,disk_name)
            with open(disk_path,'wb') as f:f.write(pdf.getvalue())
            fields=json.dumps(data,ensure_ascii=False,default=str)
            query("INSERT INTO hr_document_drafts(employee_id,document_type,document_date,reference_no,notes,fields_json,status,generated_file,created_by) VALUES(%s,'Promotion Letter',%s,%s,%s,%s,'Generated',%s,%s)",
                  (selected_employee['id'],data['document_date'],data['reference_no'],data['additional_terms'] or None,fields,disk_name,session['admin_id']),commit=True)
            query("INSERT INTO employee_documents(employee_id,title,document_type,file_name,original_name,verification_status,verification_source,verified_at) VALUES(%s,%s,'Promotion Letter',%s,%s,'Verified','HRMS Generated',NOW())",
                  (selected_employee['id'],f"Promotion Letter - {data['document_date']}",disk_name,filename),commit=True)
            notify_employee(selected_employee['id'],'New Promotion Letter','Promotion Letter has been issued and is available in your Documents section.','Document')
            audit_log('ISSUE_HR_DOCUMENT','EmployeeDocument',None,{'type':'Promotion Letter'},selected_employee['id'])
            flash('Promotion Letter generated and issued successfully. It is now available in the employee Documents section.','success')
            return redirect(url_for('admin_hr_documents',employee_id=selected_employee['id'],document_type='Promotion Letter'))
        elif dtype=='Pay Slip':
            data=pay_slip_form_data(selected_employee,company,request.form)
            background=os.path.join(app.static_folder,'images','grsj_offer_letterhead_clean.png')
            pdf=build_pay_slip_pdf(data,background)
            action=(request.form.get('action') or 'preview').strip().lower()
            safe_id=re.sub(r'[^A-Za-z0-9_-]+','_',str(selected_employee.get('login_id') or selected_employee['id']))
            filename=f"GRSJ_Pay_Slip_{safe_id}_{data['salary_month'].replace('-','')}.pdf"
            if action=='preview':
                return send_file(pdf,mimetype='application/pdf',as_attachment=False,download_name=filename)
            disk_name=f"hr_{selected_employee['id']}_{int(datetime.now().timestamp())}_{filename}"[:255]
            disk_path=os.path.join(DOCUMENT_UPLOAD,disk_name)
            with open(disk_path,'wb') as f:f.write(pdf.getvalue())
            fields=json.dumps(data,ensure_ascii=False,default=str)
            query("INSERT INTO hr_document_drafts(employee_id,document_type,document_date,reference_no,notes,fields_json,status,generated_file,created_by) VALUES(%s,'Pay Slip',%s,%s,%s,%s,'Generated',%s,%s)",
                  (selected_employee['id'],data['document_date'],data['reference_no'],data['remarks'] or None,fields,disk_name,session['admin_id']),commit=True)
            query("INSERT INTO employee_documents(employee_id,title,document_type,file_name,original_name,verification_status,verification_source,verified_at) VALUES(%s,%s,'Pay Slip',%s,%s,'Verified','HRMS Generated',NOW())",
                  (selected_employee['id'],f"Pay Slip - {data['salary_month']}",disk_name,filename),commit=True)
            notify_employee(selected_employee['id'],'New Pay Slip','Pay Slip has been issued and is available in your Documents section.','Document')
            audit_log('ISSUE_HR_DOCUMENT','EmployeeDocument',None,{'type':'Pay Slip'},selected_employee['id'])
            flash('Pay Slip generated and issued successfully. It is now available in the employee Documents section.','success')
            return redirect(url_for('admin_hr_documents',employee_id=selected_employee['id'],document_type='Pay Slip'))
        elif dtype=='Experience Letter':
            data=experience_letter_form_data(selected_employee,company,request.form)
            background=os.path.join(app.static_folder,'images','grsj_offer_letterhead_clean.png')
            pdf=build_experience_letter_pdf(data,background)
            action=(request.form.get('action') or 'preview').strip().lower()
            safe_id=re.sub(r'[^A-Za-z0-9_-]+','_',str(selected_employee.get('login_id') or selected_employee['id']))
            filename=f"GRSJ_Experience_Letter_{safe_id}_{data['document_date'].replace('-','')}.pdf"
            if action=='preview':
                return send_file(pdf,mimetype='application/pdf',as_attachment=False,download_name=filename)
            disk_name=f"hr_{selected_employee['id']}_{int(datetime.now().timestamp())}_{filename}"[:255]
            disk_path=os.path.join(DOCUMENT_UPLOAD,disk_name)
            with open(disk_path,'wb') as f:f.write(pdf.getvalue())
            fields=json.dumps(data,ensure_ascii=False,default=str)
            query("INSERT INTO hr_document_drafts(employee_id,document_type,document_date,reference_no,notes,fields_json,status,generated_file,created_by) VALUES(%s,'Experience Letter',%s,%s,%s,%s,'Generated',%s,%s)",
                  (selected_employee['id'],data['document_date'],data['reference_no'],data['additional_remarks'] or None,fields,disk_name,session['admin_id']),commit=True)
            query("INSERT INTO employee_documents(employee_id,title,document_type,file_name,original_name,verification_status,verification_source,verified_at) VALUES(%s,%s,'Experience Letter',%s,%s,'Verified','HRMS Generated',NOW())",
                  (selected_employee['id'],f"Experience Letter - {data['document_date']}",disk_name,filename),commit=True)
            notify_employee(selected_employee['id'],'New Experience Letter','Experience Letter has been issued and is available in your Documents section.','Document')
            audit_log('ISSUE_HR_DOCUMENT','EmployeeDocument',None,{'type':'Experience Letter'},selected_employee['id'])
            flash('Experience Letter generated and issued successfully. It is now available in the employee Documents section.','success')
            return redirect(url_for('admin_hr_documents',employee_id=selected_employee['id'],document_type='Experience Letter'))
        elif dtype=='Performance Improvement Plan (PIP)':
            data=pip_form_data(selected_employee,company,request.form)
            required=['period_start','period_end','review_date','performance_concerns','expectations','action_plan']
            if any(not data.get(k) for k in required):
                flash('Complete the required PIP dates, concerns, expectations and action plan.','warning')
                return redirect(url_for('admin_hr_documents',employee_id=selected_employee['id'],document_type=dtype))
            background=os.path.join(app.static_folder,'images','grsj_offer_letterhead_clean.png')
            pdf=build_pip_letter_pdf(data,background)
            action=(request.form.get('action') or 'preview').strip().lower()
            filename=f"PIP_{selected_employee['login_id']}_{data['document_date'].replace('-','')}.pdf"
            if action=='preview': return send_file(pdf,mimetype='application/pdf',as_attachment=False,download_name=filename)
            disk_name=f"hr_{selected_employee['id']}_{int(datetime.now().timestamp())}_{filename}"[:255]
            with open(os.path.join(DOCUMENT_UPLOAD,disk_name),'wb') as fh: fh.write(pdf.getvalue())
            fields=json.dumps(data,ensure_ascii=False,default=str)
            query("INSERT INTO hr_document_drafts(employee_id,document_type,document_date,reference_no,notes,fields_json,status,generated_file,created_by) VALUES(%s,%s,%s,%s,%s,%s,'Generated',%s,%s)",
                  (selected_employee['id'],dtype,data['document_date'],data['reference_no'],data['additional_remarks'] or None,fields,disk_name,session['admin_id']),commit=True)
            query("INSERT INTO employee_documents(employee_id,title,document_type,file_name,original_name,verification_status,verification_source,verified_at) VALUES(%s,%s,%s,%s,%s,'Verified','HRMS Generated',NOW())",
                  (selected_employee['id'],f"Performance Improvement Plan - {data['document_date']}",dtype,disk_name,filename),commit=True)
            notify_employee(selected_employee['id'],'Performance Improvement Plan issued','A Performance Improvement Plan (PIP) has been issued and is available in your Documents section.','Document')
            audit_log('ISSUE_HR_DOCUMENT','EmployeeDocument',None,{'type':dtype},selected_employee['id'])
            flash('Performance Improvement Plan generated and issued successfully.','success')
            return redirect(url_for('admin_hr_documents',employee_id=selected_employee['id'],document_type=dtype))
        elif dtype=='Warning Letter':
            data=warning_letter_form_data(selected_employee,company,request.form)
            background=os.path.join(app.static_folder,'images','grsj_offer_letterhead_clean.png')
            pdf=build_warning_letter_pdf(data,background);action=(request.form.get('action') or 'preview').strip().lower()
            safe_id=re.sub(r'[^A-Za-z0-9_-]+','_',str(selected_employee.get('login_id') or selected_employee['id']))
            filename=f"GRSJ_Warning_Letter_{safe_id}_{data['document_date'].replace('-','')}.pdf"
            if action=='preview':return send_file(pdf,mimetype='application/pdf',as_attachment=False,download_name=filename)
            disk_name=f"hr_{selected_employee['id']}_{int(datetime.now().timestamp())}_{filename}"[:255]
            with open(os.path.join(DOCUMENT_UPLOAD,disk_name),'wb') as f:f.write(pdf.getvalue())
            fields=json.dumps(data,ensure_ascii=False,default=str)
            query("INSERT INTO hr_document_drafts(employee_id,document_type,document_date,reference_no,notes,fields_json,status,generated_file,created_by) VALUES(%s,'Warning Letter',%s,%s,%s,%s,'Generated',%s,%s)",
              (selected_employee['id'],data['document_date'],data['reference_no'],data['warning_subject'],fields,disk_name,session['admin_id']),commit=True)
            query("INSERT INTO employee_documents(employee_id,title,document_type,file_name,original_name,verification_status,verification_source,verified_at) VALUES(%s,%s,'Warning Letter',%s,%s,'Verified','HRMS Generated',NOW())",
              (selected_employee['id'],f"Warning Letter - {data['document_date']}",disk_name,filename),commit=True)
            notify_employee(selected_employee['id'],'New Warning Letter','Warning Letter has been issued and is available in your Documents section.','Document')
            audit_log('ISSUE_HR_DOCUMENT','EmployeeDocument',None,{'type':'Warning Letter'},selected_employee['id'])
            flash('Warning Letter generated and issued successfully. It is now available in the employee Documents section.','success')
            return redirect(url_for('admin_hr_documents',employee_id=selected_employee['id'],document_type='Warning Letter'))
        elif dtype=='Relieving Letter':
            data=relieving_letter_form_data(selected_employee,company,request.form)
            background=os.path.join(app.static_folder,'images','grsj_offer_letterhead_clean.png')
            pdf=build_relieving_letter_pdf(data,background);action=(request.form.get('action') or 'preview').strip().lower()
            safe_id=re.sub(r'[^A-Za-z0-9_-]+','_',str(selected_employee.get('login_id') or selected_employee['id']))
            filename=f"GRSJ_Relieving_Letter_{safe_id}_{data['document_date'].replace('-','')}.pdf"
            if action=='preview':return send_file(pdf,mimetype='application/pdf',as_attachment=False,download_name=filename)
            disk_name=f"hr_{selected_employee['id']}_{int(datetime.now().timestamp())}_{filename}"[:255]
            with open(os.path.join(DOCUMENT_UPLOAD,disk_name),'wb') as f:f.write(pdf.getvalue())
            fields=json.dumps(data,ensure_ascii=False,default=str)
            query("INSERT INTO hr_document_drafts(employee_id,document_type,document_date,reference_no,notes,fields_json,status,generated_file,created_by) VALUES(%s,'Relieving Letter',%s,%s,%s,%s,'Generated',%s,%s)",(selected_employee['id'],data['document_date'],data['reference_no'],data['additional_remarks'] or None,fields,disk_name,session['admin_id']),commit=True)
            query("INSERT INTO employee_documents(employee_id,title,document_type,file_name,original_name,verification_status,verification_source,verified_at) VALUES(%s,%s,'Relieving Letter',%s,%s,'Verified','HRMS Generated',NOW())",(selected_employee['id'],f"Relieving Letter - {data['document_date']}",disk_name,filename),commit=True)
            notify_employee(selected_employee['id'],'New Relieving Letter','Relieving Letter has been issued and is available in your Documents section.','Document')
            audit_log('ISSUE_HR_DOCUMENT','EmployeeDocument',None,{'type':'Relieving Letter'},selected_employee['id'])
            flash('Relieving Letter generated and issued successfully. It is now available in the employee Documents section.','success')
            return redirect(url_for('admin_hr_documents',employee_id=selected_employee['id'],document_type='Relieving Letter'))
        else:
            document_date=request.form.get('document_date') or date.today().isoformat()
            reference_no=(request.form.get('reference_no') or '').strip()
            notes=(request.form.get('notes') or '').strip()
            query("""INSERT INTO hr_document_drafts(employee_id,document_type,document_date,reference_no,notes,fields_json,status,created_by)
                     VALUES(%s,%s,%s,%s,%s,%s,'Awaiting Template',%s)""",
                  (selected_employee['id'],dtype,document_date,reference_no or None,notes or None,json.dumps({},ensure_ascii=False),session['admin_id']),commit=True)
            flash(f'{dtype} request saved. Its final company template is not enabled yet.','info')
            return redirect(url_for('admin_hr_documents',employee_id=selected_employee['id'],document_type=dtype))
    # Backfill lifecycle metadata for older generated/uploaded documents.
    query("""UPDATE employee_documents
              SET document_number=COALESCE(NULLIF(document_number,''),CONCAT('GRSJ-DOC-',LPAD(id,6,'0'))),
                  issue_date=COALESCE(issue_date,DATE(uploaded_at)),
                  document_status=COALESCE(NULLIF(document_status,''),'Active')
              WHERE document_number IS NULL OR document_number='' OR issue_date IS NULL OR document_status IS NULL""",commit=True)
    query("""UPDATE employee_documents ed
              JOIN hr_document_drafts hd ON hd.employee_id=ed.employee_id AND hd.generated_file=ed.file_name
              SET ed.uploaded_by=COALESCE(ed.uploaded_by,hd.created_by),
                  ed.issue_date=COALESCE(ed.issue_date,hd.document_date),
                  hd.document_number=COALESCE(NULLIF(hd.document_number,''),ed.document_number),
                  hd.issued_at=COALESCE(hd.issued_at,hd.updated_at)
              WHERE ed.verification_source='HRMS Generated'""",commit=True)
    drafts=query("""SELECT d.*,CONCAT(e.first_name,' ',COALESCE(e.last_name,'')) employee_name,e.login_id
                    FROM hr_document_drafts d JOIN employees e ON e.id=d.employee_id ORDER BY d.created_at DESC LIMIT 50""")
    uploaded_docs=query("""SELECT ed.*,CONCAT(e.first_name,' ',COALESCE(e.last_name,'')) employee_name,e.login_id
                            FROM employee_documents ed JOIN employees e ON e.id=ed.employee_id
                            WHERE COALESCE(ed.document_status,'Active')<>'Deleted'
                            ORDER BY ed.uploaded_at DESC LIMIT 100""")
    document_types=['Offer Letter','Appointment Letter','Promotion Letter','Termination Letter','Pay Slip','Experience Letter','Performance Improvement Plan (PIP)','Warning Letter','Relieving Letter']
    return render_template('admin_hr_documents.html',employees=employees,selected_employee=selected_employee,drafts=drafts,
                           uploaded_docs=uploaded_docs,document_types=document_types,upload_document_types=EMPLOYEE_DOCUMENT_TYPES,
                           selected_type=selected_type,company=company,default_acceptance=default_acceptance)


@app.post('/admin/hr-documents/<int:draft_id>/delete')
@admin_required
def admin_hr_document_delete(draft_id):
    d=query("SELECT * FROM hr_document_drafts WHERE id=%s",(draft_id,),one=True)
    if not d:
        flash('Document record not found.','warning')
        return redirect(url_for('admin_hr_documents'))

    generated_file=(d.get('generated_file') or '').strip()
    employee_id=d.get('employee_id')
    document_type=d.get('document_type') or ''

    # Remove from active employee view but preserve database/audit history.
    if generated_file:
        ed=query("SELECT id FROM employee_documents WHERE employee_id=%s AND file_name=%s",(employee_id,generated_file),one=True)
        if ed:
            query("UPDATE employee_documents SET document_status='Deleted',deleted_at=NOW(),deleted_by=%s WHERE id=%s",
                  (session['admin_id'],ed['id']),commit=True)
        try:
            os.remove(os.path.join(DOCUMENT_UPLOAD,generated_file))
        except OSError:
            pass

    audit_log('DELETE_HR_DOCUMENT','HRDocument',draft_id,{'type':document_type,'file':generated_file},employee_id)
    query("DELETE FROM hr_document_drafts WHERE id=%s",(draft_id,),commit=True)
    notify_employee(employee_id,'Document removed',f'{document_type or "Document"} was removed from your active Documents section.','Document')
    flash(f'{document_type or "Document"} deleted successfully. Audit history was preserved.','success')
    return redirect(url_for('admin_hr_documents'))



@app.post('/admin/documents/upload')
@admin_required
def admin_employee_document_upload():
    eid=int(request.form.get('employee_id') or 0)
    dtype=(request.form.get('document_type') or 'Other').strip()
    title=(request.form.get('title') or dtype).strip()
    issue_date=(request.form.get('issue_date') or '').strip()
    expiry_date=(request.form.get('expiry_date') or '').strip()
    remarks=(request.form.get('remarks') or '').strip()
    f=request.files.get('document')
    if not eid or not f or not f.filename:
        flash('Select an employee and document file.','warning')
        return redirect(url_for('admin_hr_documents'))
    if dtype not in EMPLOYEE_DOCUMENT_TYPES:
        dtype='Other'
    ext=f.filename.rsplit('.',1)[-1].lower() if '.' in f.filename else ''
    if ext not in DOC_ALLOWED:
        flash('Unsupported document file type.','danger')
        return redirect(url_for('admin_hr_documents'))
    safe=secure_filename(f.filename)
    name=f"doc_{eid}_{int(datetime.now().timestamp())}_{secrets.token_hex(3)}_{safe}"[:255]
    f.save(os.path.join(DOCUMENT_UPLOAD,name))
    doc_no=f"GRSJ-DOC-{eid}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    did=query("""INSERT INTO employee_documents
                 (employee_id,document_number,version_no,title,document_type,file_name,original_name,
                  verification_status,verification_source,issue_date,expiry_date,document_status,remarks,uploaded_by)
                 VALUES(%s,%s,1,%s,%s,%s,%s,'Pending','Manual Upload',NULLIF(%s,''),NULLIF(%s,''),'Active',%s,%s)""",
              (eid,doc_no,title[:150],dtype[:80],name,f.filename[:255],issue_date,expiry_date,remarks[:500] or None,session['admin_id']),commit=True)
    audit_log('UPLOAD_DOCUMENT','EmployeeDocument',did,{'type':dtype,'document_number':doc_no},eid)
    notify_employee(eid,'New document added',f'{dtype} has been added to your Documents section.','Document')
    flash('Employee document uploaded successfully.','success')
    return redirect(url_for('admin_hr_documents'))

@app.post('/admin/documents/<int:did>/delete')
@admin_required
def admin_employee_document_delete(did):
    d=query("SELECT * FROM employee_documents WHERE id=%s",(did,),one=True)
    if not d:
        flash('Document not found.','warning')
        return redirect(url_for('admin_hr_documents'))
    query("UPDATE employee_documents SET document_status='Deleted',deleted_at=NOW(),deleted_by=%s WHERE id=%s",
          (session['admin_id'],did),commit=True)
    audit_log('DELETE_DOCUMENT','EmployeeDocument',did,{'title':d.get('title'),'file':d.get('file_name')},d.get('employee_id'))
    notify_employee(d['employee_id'],'Document removed',f'{d.get("title") or "A document"} was removed from your active Documents list.','Document')
    flash('Document removed from the active employee record. Audit history is preserved.','success')
    return redirect(url_for('admin_hr_documents'))

@app.post('/employee/documents/<int:did>/acknowledge')
@employee_required
def employee_document_acknowledge(did):
    eid=session['employee_id']
    d=query("SELECT id FROM employee_documents WHERE id=%s AND employee_id=%s AND COALESCE(document_status,'Active')<>'Deleted'",(did,eid),one=True)
    if d:
        query("UPDATE employee_documents SET acknowledged_at=COALESCE(acknowledged_at,NOW()) WHERE id=%s",(did,),commit=True)
    return redirect(url_for('employee_documents'))

@app.route('/employee/notifications')
@employee_required
def employee_notifications():
    eid=session['employee_id']
    rows=query("SELECT * FROM employee_notifications WHERE employee_id=%s ORDER BY created_at DESC LIMIT 100",(eid,))
    query("UPDATE employee_notifications SET is_read=1,read_at=COALESCE(read_at,NOW()) WHERE employee_id=%s AND is_read=0",(eid,),commit=True)
    return render_template('employee_notifications.html',notifications=rows)

@app.route('/employee/profile')
@employee_required
def employee_profile():
    eid=session['employee_id'];emp=query("SELECT e.*,d.department_name FROM employees e LEFT JOIN departments d ON d.id=e.department_id WHERE e.id=%s",(eid,),one=True);return render_template('employee_profile.html',employee=emp,profile_completion=profile_completion(emp))
@app.route('/employee/company-details')
@employee_required
def employee_company_details():return render_template('employee_company_details.html',details=settings())
@app.route('/employee/holidays')
@employee_required
def employee_holidays():rows=query("SELECT * FROM holidays WHERE status='Active' AND holiday_date>=CURDATE() ORDER BY holiday_date");return render_template('employee_holidays.html',holidays=rows)
@app.route('/employee/documents')
@employee_required
def employee_documents():
    eid=session['employee_id']
    rows=query("""SELECT *,CASE WHEN expiry_date IS NOT NULL AND expiry_date<CURDATE() THEN 'Expired'
                               WHEN expiry_date IS NOT NULL AND expiry_date<=DATE_ADD(CURDATE(),INTERVAL 30 DAY) THEN 'Expiring Soon'
                               ELSE COALESCE(document_status,'Active') END lifecycle_status
                  FROM employee_documents
                  WHERE employee_id=%s AND COALESCE(document_status,'Active')<>'Deleted'
                  ORDER BY uploaded_at DESC""",(eid,))
    return render_template('employee_documents.html',documents=rows)


@app.route('/employee/allocation')
@employee_required
def employee_allocation():
    eid=session['employee_id']
    emp=query("SELECT e.*,d.department_name FROM employees e LEFT JOIN departments d ON d.id=e.department_id WHERE e.id=%s",(eid,),one=True)
    snap=query("""SELECT * FROM allocation_snapshots
                  WHERE employee_id=%s ORDER BY imported_at DESC,id DESC LIMIT 1""",(eid,),one=True)
    headers=[]; rows=[]
    if snap:
        try:
            headers=json.loads(snap.get('headers_json') or '[]')
            rows=json.loads(snap.get('rows_json') or '[]')
        except (ValueError,TypeError,json.JSONDecodeError):
            headers=[]; rows=[]
    display_rows=[]
    if rows:
        total_index=None
        last=rows[-1] if rows else []
        if len(last)>=2 and not str(last[0] or '').strip() and not str(last[1] or '').strip() and any(str(x or '').strip() for x in last[2:]):
            total_index=len(rows)-1

        i=0
        while i < len(rows):
            row=rows[i]

            if total_index is not None and i == total_index:
                display_rows.append({'cells':row,'is_total':True,'show_ro':False,'ro_rowspan':0})
                i += 1
                continue

            ro=str(row[0] or '').strip() if row else ''
            if ro:
                span=1
                j=i+1
                while j < len(rows):
                    if total_index is not None and j == total_index:
                        break
                    next_row=rows[j]
                    next_ro=str(next_row[0] or '').strip() if next_row else ''
                    if next_ro:
                        break
                    span += 1
                    j += 1

                display_rows.append({'cells':row,'is_total':False,'show_ro':True,'ro_rowspan':span})
                for k in range(i+1,j):
                    display_rows.append({'cells':rows[k],'is_total':False,'show_ro':False,'ro_rowspan':0})
                i=j
            else:
                display_rows.append({'cells':row,'is_total':False,'show_ro':False,'ro_rowspan':0})
                i += 1

    return render_template('employee_allocation.html',employee=emp,snapshot=snap,headers=headers,rows=rows,display_rows=display_rows)


@app.post('/employee/allocation/refresh')
@employee_required
def employee_allocation_refresh():
    eid=session['employee_id']
    try:
        from allocation_sync import sync_employee
        sync_employee(eid,force=True)
        flash('Performance Tracking refreshed from your latest Koofr workbook.','success')
    except Exception as e:
        audit_log('EMPLOYEE_PERFORMANCE_REFRESH_FAILED','Performance',eid,{'error':str(e)},eid)
        flash('Performance data could not be refreshed right now. The administrator can review the sync log.','danger')
    return redirect(url_for('employee_allocation'))

@app.route('/employee/leaves',methods=['GET','POST'])
@employee_required
def employee_leaves():
    eid=session['employee_id']
    if request.method=='POST':
        f=request.form;sd=datetime.strptime(f.get('start_date'),'%Y-%m-%d').date();ed=datetime.strptime(f.get('end_date'),'%Y-%m-%d').date();days=(ed-sd).days+1;days=.5 if f.get('day_type')!='Full Day' else days;reason=f.get('reason','').strip();contact=f.get('contact_during_leave','').strip()
        leave_type_id=f.get('leave_type_id'); leave_type=query("SELECT name FROM leave_types WHERE id=%s AND status='Active'",(leave_type_id,),one=True)
        leave_name=(leave_type.get('name') if leave_type else '').strip(); is_custom=('custom' in leave_name.lower() or leave_name.lower()=='other')
        custom=(f.get('custom_leave_type') or '').strip() if is_custom else ''
        if custom:reason=f"[{custom}] {reason}"
        if not leave_type:flash('Please select a valid leave type.','danger')
        elif is_custom and not custom:flash('Please enter the custom leave type.','danger')
        elif ed<sd:flash('End date cannot be before start date.','danger')
        elif contact and (len(contact)!=10 or not contact.isdigit()):flash('Contact during leave must be exactly 10 numeric digits.','danger')
        elif len(reason)>500:flash('Leave reason is too long. Please keep it within 500 characters.','danger')
        else:query("INSERT INTO leave_requests(employee_id,leave_type_id,start_date,end_date,day_type,total_days,reason,contact_during_leave,status,admin_remark) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,'Pending',NULL)",(eid,leave_type_id,sd,ed,f.get('day_type'),days,reason,contact or None),commit=True);flash('Leave request submitted for administrator review.','success')
        return redirect(url_for('employee_leaves'))
    types=query("SELECT * FROM leave_types WHERE status='Active' ORDER BY name");rows=query("SELECT lr.*,lt.name leave_type FROM leave_requests lr JOIN leave_types lt ON lt.id=lr.leave_type_id WHERE employee_id=%s ORDER BY created_at DESC",(eid,));return render_template('employee_leaves.html',types=types,rows=rows)


# ===== v1.0.28 Operations & Governance =====
def _escalate_pending_approvals():
    """SLA expiry moves HR -> Manager -> Owner. Owner is terminal and receives reminders."""
    try:
        rows=query("SELECT * FROM approval_requests WHERE status='Pending' AND current_stage IN ('HR','Manager','Owner') AND TIMESTAMPDIFF(HOUR,COALESCE(stage_entered_at,created_at),NOW())>=COALESCE(sla_hours,24)")
        for r in rows:
            stage=r['current_stage']
            if stage=='HR':
                query("UPDATE approval_requests SET current_stage='Manager',stage_entered_at=NOW(),escalation_count=escalation_count+1 WHERE id=%s",(r['id'],),commit=True)
                query("INSERT INTO admin_notifications(admin_role,title,message,entity_type,entity_id) VALUES('Manager','Auto-escalated approval',%s,'ApprovalRequest',%s)",(f"{r['title']} was not actioned by HR within SLA.",str(r['id'])),commit=True)
                audit_log('AUTO_ESCALATE_HR_MANAGER','ApprovalRequest',r['id'],{'reason':'SLA expired'})
            elif stage=='Manager':
                query("UPDATE approval_requests SET current_stage='Owner',stage_entered_at=NOW(),escalation_count=escalation_count+1 WHERE id=%s",(r['id'],),commit=True)
                query("INSERT INTO admin_notifications(admin_role,title,message,entity_type,entity_id) VALUES('Owner','Approval escalated to Owner',%s,'ApprovalRequest',%s)",(f"{r['title']} was not actioned by Manager within SLA.",str(r['id'])),commit=True)
                audit_log('AUTO_ESCALATE_MANAGER_OWNER','ApprovalRequest',r['id'],{'reason':'SLA expired'})
            elif stage=='Owner' and (not r.get('last_reminder_at') or (datetime.now()-r['last_reminder_at']).total_seconds()>=86400):
                query("INSERT INTO admin_notifications(admin_role,title,message,entity_type,entity_id) VALUES('Owner','Owner approval still pending',%s,'ApprovalRequest',%s)",(f"Final action is still required for {r['title']}.",str(r['id'])),commit=True)
                query("UPDATE approval_requests SET last_reminder_at=NOW() WHERE id=%s",(r['id'],),commit=True)
    except Exception: pass

@app.before_request
def governance_tick():
    if request.path.startswith('/static/') or request.path.startswith('/uploads/'): return
    now=int(datetime.now().timestamp())
    if now-int(app.config.get('_last_governance_tick',0))>=300:
        app.config['_last_governance_tick']=now
        _escalate_pending_approvals()
        try:
            fn=app.config.get('_automation_tick')
            if fn: fn()
        except Exception: pass

@app.route('/admin/action-center')
@admin_required
def admin_action_center():
    role=current_admin_role()
    where="status='Pending'"; params=[]
    if role!='Super Admin': where+=" AND current_stage=%s"; params=[role]
    approvals=query(f"SELECT * FROM approval_requests WHERE {where} ORDER BY created_at",tuple(params))
    regs=query("SELECT r.*,CONCAT(e.first_name,' ',COALESCE(e.last_name,'')) employee_name FROM attendance_regularization r JOIN employees e ON e.id=r.employee_id WHERE r.status='Pending' ORDER BY r.created_at")
    open_cases=query("SELECT COUNT(*) c FROM hr_cases WHERE status<>'Closed'",one=True)
    pending_exit=query("SELECT COUNT(*) c FROM employee_exit_checklist x JOIN employees e ON e.id=x.employee_id WHERE e.status='Inactive' AND (x.asset_returned=0 OR x.handover_completed=0 OR x.documents_completed=0 OR x.attendance_closed=0)",one=True)
    return render_template('admin_action_center.html',approvals=approvals,regularizations=regs,open_cases=(open_cases or {}).get('c',0),pending_exit=(pending_exit or {}).get('c',0))

@app.get('/api/employee/attendance-status')
@employee_required
def employee_attendance_status():
    eid = session['employee_id']
    date_str = (request.args.get('date') or '').strip()

    if not date_str:
        return jsonify({'ok': False, 'message': 'Date is required.'}), 400

    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'ok': False, 'message': 'Invalid date.'}), 400

    attendance = query(
        """
        SELECT attendance_date, check_in, check_out, status, remarks
        FROM attendance
        WHERE employee_id=%s AND attendance_date=%s
        LIMIT 1
        """,
        (eid, selected_date),
        one=True
    )

    if not attendance:
        return jsonify({
            'ok': True,
            'found': False,
            'date': date_str,
            'status': 'No Attendance Record',
            'check_in': None,
            'check_out': None,
            'remarks': None
        })

    def format_dt(value):
        if not value:
            return None
        if hasattr(value, 'strftime'):
            return value.strftime('%H:%M')
        return str(value)[:5]

    return jsonify({
        'ok': True,
        'found': True,
        'date': date_str,
        'status': attendance.get('status') or 'Unknown',
        'check_in': format_dt(attendance.get('check_in')),
        'check_out': format_dt(attendance.get('check_out')),
        'remarks': attendance.get('remarks') or ''
    })

@app.route('/employee/attendance-regularization',methods=['GET','POST'])
@employee_required
def employee_regularization():
    eid=session['employee_id']
    if request.method=='POST':
        f=request.form; d=f.get('attendance_date'); typ=f.get('request_type'); reason=(f.get('reason') or '').strip()

        requested_status = f.get('requested_status','').strip()

        if not d or not typ or not requested_status or len(reason) < 3: flash('Date, request type and reason are required.','danger')

        allowed_statuses = {'Present', 'Late', 'Half Day', 'Absent'}

        if requested_status not in allowed_statuses:
            flash('Invalid attendance status selected.','danger')
        else:
            rid=query("""
                INSERT INTO attendance_regularization(
                    employee_id,
                    attendance_date,
                    request_type,
                    requested_check_in,
                    requested_check_out,
                    requested_status,
                    reason
                )
                VALUES(
                    %s,
                    %s,
                    %s,
                    NULLIF(%s,''),
                    NULLIF(%s,''),
                    NULLIF(%s,''),
                    %s
                )
            """,(
                eid,
                d,
                typ,
                f.get('requested_check_in',''),
                f.get('requested_check_out',''),
                f.get('requested_status',''),
                reason[:500]
            ),commit=True)
            ref=f"ATT-{datetime.now():%Y}-{int(rid):04d}"
            system_admin=query("SELECT id FROM admins WHERE is_active=1 ORDER BY FIELD(role,'Super Admin','Owner','Manager','HR'),id LIMIT 1",one=True); aid=query("INSERT INTO approval_requests(request_type,entity_type,entity_id,title,payload_json,status,current_stage,submitted_by,stage_entered_at,sla_hours,reference_no) VALUES('Attendance Regularization','AttendanceRegularization',%s,%s,%s,'Pending','HR',%s,NOW(),24,%s)",(str(rid),f"Attendance regularization · {d}",json.dumps({'regularization_id':rid}),system_admin['id'],ref),commit=True)
            query("UPDATE attendance_regularization SET approval_request_id=%s WHERE id=%s",(aid,rid),commit=True); flash(f'Request submitted. Reference: {ref}','success')
        return redirect(url_for('employee_regularization'))
    rows=query("SELECT * FROM attendance_regularization WHERE employee_id=%s ORDER BY created_at DESC",(eid,)); return render_template('employee_regularization.html',rows=rows)

@app.route('/admin/assets',methods=['GET','POST'])
@admin_required
def admin_assets():
    if request.method=='POST':
        f=request.form
        query("INSERT INTO employee_assets(employee_id,asset_type,asset_name,asset_code,serial_no,issue_date,issue_condition,status,remarks,created_by) VALUES(%s,%s,%s,%s,%s,NULLIF(%s,''),%s,'Issued',%s,%s)",(f.get('employee_id'),f.get('asset_type'),f.get('asset_name'),f.get('asset_code') or None,f.get('serial_no') or None,f.get('issue_date',''),f.get('issue_condition') or None,f.get('remarks') or None,session['admin_id']),commit=True)
        audit_log('ISSUE_ASSET','Asset',None,{'employee_id':f.get('employee_id'),'asset':f.get('asset_name')}); flash('Asset issued and recorded.','success'); return redirect(url_for('admin_assets'))
    rows=query("SELECT a.*,CONCAT(e.first_name,' ',COALESCE(e.last_name,'')) employee_name FROM employee_assets a JOIN employees e ON e.id=a.employee_id ORDER BY a.created_at DESC"); emps=query("SELECT id,first_name,last_name,employee_id FROM employees ORDER BY first_name")
    return render_template('admin_assets.html',rows=rows,employees=emps)

@app.post('/admin/assets/<int:aid>/return')
@admin_required
def admin_asset_return(aid):
    query("UPDATE employee_assets SET status='Returned',return_date=CURDATE(),return_condition=%s WHERE id=%s",((request.form.get('return_condition') or '')[:120],aid),commit=True); audit_log('RETURN_ASSET','Asset',aid); flash('Asset marked returned.','success'); return redirect(url_for('admin_assets'))

@app.route('/admin/announcements',methods=['GET','POST'])
@admin_required
def admin_announcements():
    if request.method=='POST':
        f=request.form; query("INSERT INTO company_announcements(title,message,announcement_type,priority,publish_from,publish_until,requires_ack,status,created_by) VALUES(%s,%s,%s,%s,NOW(),NULLIF(%s,''),%s,'Active',%s)",((f.get('title') or '')[:180],f.get('message'),f.get('announcement_type','Notice'),f.get('priority','Normal'),f.get('publish_until',''),1 if f.get('requires_ack') else 0,session['admin_id']),commit=True); audit_log('PUBLISH_ANNOUNCEMENT','Announcement'); flash('Announcement published.','success'); return redirect(url_for('admin_announcements'))
    rows=query("SELECT a.*,(SELECT COUNT(*) FROM announcement_acknowledgements x WHERE x.announcement_id=a.id) ack_count FROM company_announcements a ORDER BY created_at DESC"); return render_template('admin_announcements.html',rows=rows)

@app.route('/employee/announcements')
@employee_required
def employee_announcements():
    rows=query("SELECT a.*,x.acknowledged_at FROM company_announcements a LEFT JOIN announcement_acknowledgements x ON x.announcement_id=a.id AND x.employee_id=%s WHERE a.status='Active' AND a.publish_from<=NOW() AND (a.publish_until IS NULL OR a.publish_until>=NOW()) ORDER BY FIELD(a.priority,'Critical','High','Normal'),a.created_at DESC",(session['employee_id'],)); return render_template('employee_announcements.html',rows=rows)

@app.post('/employee/announcements/<int:aid>/ack')
@employee_required
def employee_announcement_ack(aid):
    query("INSERT IGNORE INTO announcement_acknowledgements(announcement_id,employee_id) VALUES(%s,%s)",(aid,session['employee_id']),commit=True); return redirect(url_for('employee_announcements'))

@app.route('/admin/cases',methods=['GET','POST'])
@admin_required
def admin_cases():
    if current_admin_role() not in ('HR','Owner','Super Admin'):
        flash('HR cases are confidential.','danger'); return redirect(url_for('admin_dashboard'))
    if request.method=='POST':
        f=request.form; seq=query("SELECT COALESCE(MAX(id),0)+1 n FROM hr_cases",one=True)['n']; case_no=f"CASE-{datetime.now():%Y}-{int(seq):04d}"
        query("INSERT INTO hr_cases(case_no,employee_id,case_type,title,description,severity,opened_by) VALUES(%s,%s,%s,%s,%s,%s,%s)",(case_no,f.get('employee_id'),f.get('case_type'),f.get('title'),f.get('description'),f.get('severity','Normal'),session['admin_id']),commit=True); audit_log('OPEN_HR_CASE','HRCase',None,{'case_no':case_no}); flash(f'{case_no} opened.','success'); return redirect(url_for('admin_cases'))
    rows=query("SELECT c.*,CONCAT(e.first_name,' ',COALESCE(e.last_name,'')) employee_name FROM hr_cases c JOIN employees e ON e.id=c.employee_id ORDER BY c.created_at DESC"); emps=query("SELECT id,first_name,last_name,employee_id FROM employees ORDER BY first_name"); return render_template('admin_cases.html',rows=rows,employees=emps)

@app.post('/admin/cases/<int:cid>/close')
@admin_required
def admin_case_close(cid):
    if current_admin_role() not in ('Owner','Super Admin'): flash('Only Owner or Super Admin can close a case.','danger'); return redirect(url_for('admin_cases'))
    query("UPDATE hr_cases SET status='Closed',resolution=%s,closed_by=%s,closed_at=NOW() WHERE id=%s",((request.form.get('resolution') or '')[:2000],session['admin_id'],cid),commit=True); audit_log('CLOSE_HR_CASE','HRCase',cid); return redirect(url_for('admin_cases'))

@app.route('/admin/employee/<int:eid>/lifecycle',methods=['GET','POST'])
@admin_required
def admin_employee_lifecycle(eid):
    emp=query("SELECT * FROM employees WHERE id=%s",(eid,),one=True)
    if not emp: return ('Not found',404)
    if request.method=='POST':
        f=request.form; query("INSERT INTO employee_lifecycle(employee_id,event_type,title,details,effective_date,created_by) VALUES(%s,%s,%s,%s,%s,%s)",(eid,f.get('event_type'),f.get('title'),f.get('details') or None,f.get('effective_date'),session['admin_id']),commit=True); audit_log('ADD_LIFECYCLE_EVENT','Employee',eid,{'event_type':f.get('event_type')},eid); return redirect(url_for('admin_employee_lifecycle',eid=eid))
    rows=query("SELECT l.*,a.full_name admin_name FROM employee_lifecycle l LEFT JOIN admins a ON a.id=l.created_by WHERE l.employee_id=%s ORDER BY l.effective_date DESC,l.id DESC",(eid,)); return render_template('admin_employee_lifecycle.html',employee=emp,rows=rows)

@app.route('/admin/onboarding/<int:eid>',methods=['GET','POST'])
@admin_required
def admin_onboarding(eid):
    emp=query("SELECT * FROM employees WHERE id=%s",(eid,),one=True)
    if request.method=='POST':
        f=request.form; vals=[1 if f.get(x) else 0 for x in ('documents_verified','id_card_issued','policy_acknowledged','assets_issued','orientation_completed')]
        query("INSERT INTO onboarding_checklist(employee_id,documents_verified,id_card_issued,policy_acknowledged,assets_issued,orientation_completed,remarks,updated_by) VALUES(%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE documents_verified=VALUES(documents_verified),id_card_issued=VALUES(id_card_issued),policy_acknowledged=VALUES(policy_acknowledged),assets_issued=VALUES(assets_issued),orientation_completed=VALUES(orientation_completed),remarks=VALUES(remarks),updated_by=VALUES(updated_by)",(eid,*vals,(f.get('remarks') or '')[:500],session['admin_id']),commit=True); audit_log('UPDATE_ONBOARDING','Employee',eid,None,eid); flash('Onboarding checklist updated.','success'); return redirect(url_for('admin_onboarding',eid=eid))
    row=query("SELECT * FROM onboarding_checklist WHERE employee_id=%s",(eid,),one=True) or {}; return render_template('admin_onboarding.html',employee=emp,item=row)

@app.route('/admin/notifications')
@admin_required
def admin_notifications():
    role=current_admin_role(); rows=query("SELECT * FROM admin_notifications WHERE admin_role IN (%s,'All') ORDER BY created_at DESC LIMIT 100",(role,)); query("UPDATE admin_notifications SET is_read=1 WHERE admin_role IN (%s,'All')",(role,),commit=True); return render_template('admin_notifications.html',rows=rows)

@app.errorhandler(404)
def notfound(e):return render_template('error.html',code=404,message='Page not found'),404
@app.errorhandler(500)
def servererr(e):return render_template('error.html',code=500,message='Something went wrong. Check the terminal for details.'),500

from advanced_features import register_advanced_features
register_advanced_features(app, query, admin_required, role_required, current_admin_role, audit_log, settings)

if __name__=='__main__':app.run(host='127.0.0.1',port=5000,debug=os.getenv('FLASK_DEBUG','1')=='1')
